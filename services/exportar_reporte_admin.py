"""
services/exportar_reporte_admin.py
Genera el Reporte de Checador para personal administrativo,
clon del formato producido por el sistema v1.

Estructura del Excel (hoja REPORTE):
  Fila 1 : "REPORTE CHECADOR 2026"  A1:Q1 merge
  Fila 2 : "Fecha actual: YYYY-MM-DD - YYYY-MM-DD"  A2:Q2 merge
  Fila 3 : vacía

  Por cada trabajador — bloque de 8 filas (base = 4 + índice*8):
    base+0  : Encabezados de día (A-N) + INCIDENCIAS / OBSERVACIONES / FIRMA (O-Q)
    base+1  : ID:  {chec_id}    Nombre:  {nombre completo}
    base+2  : Horas de ENTRADA  (rosa en celdas con retardo, "F" si falta)
    base+3  : Horas de SALIDA A COMER  (verde si tiene_comida)
    base+4  : Horas de REGRESO DE COMER  (verde si tiene_comida)
    base+5  : Horas de SALIDA
    base+6  : vacía
    base+7  : vacía
    Columnas O / P / Q mergeadas sobre base+2 : base+5

Colores:
  Verde  (#8DEB8E) — celdas de comida (filas 3 y 4 del bloque) con valor
  Rosa   (#F4CCCC) — celda de entrada con retardo (> 10 min)
  Rojo   (#FF0000) — fuente del encabezado del día domingo

Checadas de comida:
  Se intenta primero con tipo_punch=2 (break-out) y tipo_punch=3 (break-in).
  Si el dispositivo no emite esos tipos, se usa heurística de ventana ±45 min
  sobre el punto medio de la jornada.
"""

from io import BytesIO
from datetime import date, time, timedelta
from collections import defaultdict
from typing import Optional

import psycopg2
from psycopg2.extras import RealDictCursor
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from config import settings

# ── Constantes de estilo ────────────────────────────────────────────────────────

_THIN        = Side(style="thin")
BORDER_THIN  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
FILL_GREEN   = PatternFill("solid", fgColor="8DEB8E")
FILL_PINK    = PatternFill("solid", fgColor="F4CCCC")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center")
ALIGN_WRAP   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

FONT_BOLD    = Font(name="Calibri", size=11, bold=True)
FONT_NORMAL  = Font(name="Calibri", size=11)
FONT_RED     = Font(name="Calibri", size=11, bold=True, color="FF0000")

# Abreviatura de día (Python weekday 0=lun … 6=dom)
DAY_ABBR = {0: "L", 1: "M", 2: "M", 3: "J", 4: "V", 5: "S", 6: "D"}
DAY_COL  = {
    0: "lunes", 1: "martes", 2: "miercoles", 3: "jueves",
    4: "viernes", 5: "sabado", 6: "domingo",
}

TOLERANCIA_COMIDA_SEG = 90 * 60   # ±90 min alrededor del punto medio de la jornada


# ── Helpers ─────────────────────────────────────────────────────────────────────

def _get_conn():
    return psycopg2.connect(settings.database_url_nomina, cursor_factory=RealDictCursor)


def _td_to_time(t) -> time:
    """Convierte timedelta o time a datetime.time."""
    if isinstance(t, timedelta):
        total = int(t.total_seconds())
        h, rem = divmod(total, 3600)
        m, s   = divmod(rem, 60)
        return time(h % 24, m, s)
    return t


def _time_to_secs(t) -> int:
    if isinstance(t, timedelta):
        return int(t.total_seconds())
    return t.hour * 3600 + t.minute * 60 + t.second


def _fmt_hms(t) -> str:
    """Devuelve 'HH:MM:SS' desde time o timedelta."""
    tt = _td_to_time(t)
    return f"{tt.hour:02d}:{tt.minute:02d}:{tt.second:02d}"


def _dedup(horas: list, gap_secs: int = 180) -> list:
    """Elimina checadas duplicadas dentro de gap_secs segundos."""
    if not horas:
        return []
    result = [horas[0]]
    for h in horas[1:]:
        if _time_to_secs(h) - _time_to_secs(result[-1]) >= gap_secs:
            result.append(h)
    return result


def _cell_border(ws, row: int, col: int):
    ws.cell(row=row, column=col).border = BORDER_THIN


def _apply_borders_row(ws, row: int, min_col: int = 1, max_col: int = 17):
    for c in range(min_col, max_col + 1):
        _cell_border(ws, row, c)


def _clasificar_checadas(
    checadas_raw: list,
    hora_entrada: time,
    hora_salida: time,
    tiene_comida: bool,
) -> dict:
    """
    Dada la lista de checadas del día (dicts con 'hora' y 'tipo_punch'),
    devuelve:
        entrada, salida, comida_sal, comida_ent  (time|None cada uno)
        es_retardo  (bool)
        comida_incompleta (bool)
    """
    checadas_sorted = sorted(checadas_raw, key=lambda x: _time_to_secs(x["hora"]))

    entrada_secs = _time_to_secs(hora_entrada)
    salida_secs  = _time_to_secs(hora_salida)
    mid_secs     = (entrada_secs + salida_secs) // 2

    entrada_checada:    Optional[time] = None
    salida_checada:     Optional[time] = None
    comida_sal_checada: Optional[time] = None
    comida_ent_checada: Optional[time] = None

    # MB360 envía casi todo como tipo_punch=0 → no es confiable.
    # Algoritmo por zonas: comida se clasifica PRIMERO (±90 min del punto medio);
    # entrada/salida se asignan de lo que queda fuera de esa zona.
    all_horas = [_td_to_time(c["hora"]) for c in checadas_sorted]

    by_type: dict[int, list] = defaultdict(list)
    for c in checadas_sorted:
        by_type[c["tipo_punch"]].append(_td_to_time(c["hora"]))
    has_dedicated_comida = bool(by_type.get(2) or by_type.get(3))

    if has_dedicated_comida:
        # El dispositivo envió tipos explícitos de comida → confiar en ellos
        comida_sal_checada = by_type[2][0]  if by_type.get(2) else None
        comida_ent_checada = by_type[3][0]  if by_type.get(3) else None
        entrada_checada    = by_type[0][0]  if by_type.get(0) else None
        salida_checada     = by_type[1][-1] if by_type.get(1) else None
    elif tiene_comida:
        # Separar en zona comida (±TOLERANCIA del punto medio) y zona trabajo
        zona_comida  = [h for h in all_horas
                        if abs(_time_to_secs(h) - mid_secs) <= TOLERANCIA_COMIDA_SEG]
        zona_trabajo = [h for h in all_horas
                        if abs(_time_to_secs(h) - mid_secs) > TOLERANCIA_COMIDA_SEG]

        if zona_comida:
            comida_sal_checada = zona_comida[0]
        if len(zona_comida) >= 2:
            comida_ent_checada = zona_comida[1]

        antes   = [h for h in zona_trabajo if _time_to_secs(h) < mid_secs]
        despues = [h for h in zona_trabajo if _time_to_secs(h) > mid_secs]
        if antes:
            entrada_checada = antes[0]
        if despues:
            salida_checada = despues[-1]
    else:
        # Sin comida: primera=entrada, última=salida
        entrada_checada = all_horas[0]        if all_horas else None
        salida_checada  = all_horas[-1] if len(all_horas) >= 2 else None

    # ── Retardo
    es_retardo = False
    if entrada_checada:
        ent_secs     = _time_to_secs(entrada_checada)
        ventana_max  = entrada_secs + 10 * 60
        es_retardo   = ent_secs > ventana_max

    comida_incompleta = (
        tiene_comida and
        (comida_sal_checada is None or comida_ent_checada is None)
    )

    return {
        "entrada":          entrada_checada,
        "salida":           salida_checada,
        "comida_sal":       comida_sal_checada,
        "comida_ent":       comida_ent_checada,
        "es_retardo":       es_retardo,
        "comida_incompleta": comida_incompleta,
    }


# ── Función principal ────────────────────────────────────────────────────────────

def generar_reporte_checador_admin(periodo_id: int) -> bytes:
    """
    Genera el Excel de reporte de checador para el periodo_admin indicado.
    Devuelve los bytes del archivo .xlsx.
    """
    conn = _get_conn()
    cur  = conn.cursor()

    # 1 ── Período
    cur.execute(
        "SELECT id, nombre, fecha_inicio, fecha_fin FROM periodos_admin WHERE id = %s",
        (periodo_id,)
    )
    periodo = cur.fetchone()
    if not periodo:
        conn.close()
        raise ValueError(f"Período {periodo_id} no encontrado")

    fecha_inicio: date = periodo["fecha_inicio"]
    fecha_fin:    date = periodo["fecha_fin"]

    dias: list[date] = []
    d = fecha_inicio
    while d <= fecha_fin:
        dias.append(d)
        d += timedelta(days=1)

    # 2 ── Trabajadores activos
    cur.execute(
        "SELECT id, chec_id, nombre FROM trabajadores WHERE activo = TRUE ORDER BY nombre"
    )
    trabajadores = [dict(r) for r in cur.fetchall()]

    # 3 ── Horarios
    trab_ids = [t["id"] for t in trabajadores]
    horarios_por_trab: dict[int, list] = defaultdict(list)
    if trab_ids:
        cur.execute(
            """SELECT trabajador_id, lunes, martes, miercoles, jueves, viernes,
                      sabado, domingo, hora_entrada, hora_salida, tiene_comida
               FROM horarios_trabajador
               WHERE trabajador_id = ANY(%s)""",
            (trab_ids,)
        )
        for h in cur.fetchall():
            horarios_por_trab[h["trabajador_id"]].append(dict(h))

    # 4 ── Checadas del período
    chec_ids = [t["chec_id"] for t in trabajadores if t["chec_id"] is not None]
    # checadas_por_chec_dia[chec_id][date] = list de {hora, tipo_punch}
    checadas_por_chec_dia: dict = defaultdict(lambda: defaultdict(list))
    if chec_ids:
        cur.execute(
            """SELECT user_id,
                      timestamp_checada::date AS dia,
                      timestamp_checada::time AS hora,
                      tipo_punch
               FROM asistencias_checadas
               WHERE user_id = ANY(%s)
                 AND timestamp_checada::date >= %s
                 AND timestamp_checada::date <= %s
               ORDER BY timestamp_checada""",
            (chec_ids, fecha_inicio, fecha_fin)
        )
        for c in cur.fetchall():
            checadas_por_chec_dia[c["user_id"]][c["dia"]].append({
                "hora":       c["hora"],
                "tipo_punch": c["tipo_punch"],
            })

    conn.close()

    # 5 ── Construir Excel ────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "REPORTE"

    # Anchos de columna
    ws.column_dimensions["A"].width = 8.71
    for c in range(2, 15):
        ws.column_dimensions[get_column_letter(c)].width = 9.14
    ws.column_dimensions["O"].width = 12.0
    ws.column_dimensions["P"].width = 25.0
    ws.column_dimensions["Q"].width = 24.0

    # ── Encabezado global (filas 1-3) ────────────────────────────────────────────
    ws.merge_cells("A1:Q1")
    c1 = ws["A1"]
    c1.value     = "REPORTE CHECADOR 2026"
    c1.font      = Font(name="Calibri", size=14, bold=True)
    c1.alignment = ALIGN_CENTER

    ws.merge_cells("A2:Q2")
    c2 = ws["A2"]
    c2.value     = (f"Fecha actual: {fecha_inicio.strftime('%Y-%m-%d')} - "
                    f"{fecha_fin.strftime('%Y-%m-%d')}")
    c2.font      = Font(name="Calibri", size=11, bold=True)
    c2.alignment = ALIGN_CENTER
    # Fila 3: vacía (separador)

    # ── Bloque por trabajador ────────────────────────────────────────────────────
    for idx, trab in enumerate(trabajadores):
        base     = 4 + idx * 8   # fila de inicio (1-indexed)
        chec_id  = trab.get("chec_id")
        horarios = horarios_por_trab.get(trab["id"], [])

        # Horario efectivo por día (primer bloque del trabajador que aplique)
        horario_por_dia: dict[date, dict] = {}
        for dia in dias:
            col = DAY_COL[dia.weekday()]
            for h in horarios:
                if h.get(col):
                    horario_por_dia[dia] = h
                    break

        # Checadas del trabajador (con dedup 3 min por día)
        raw_por_dia: dict = checadas_por_chec_dia.get(chec_id, {}) if chec_id else {}
        checadas_por_dia: dict[date, list] = {}
        for dia, raw_list in raw_por_dia.items():
            sorted_list = sorted(raw_list, key=lambda x: _time_to_secs(x["hora"]))
            # Dedup manteniendo tipo_punch: quedarme con el primero dentro de cada ventana
            deduped = []
            for c in sorted_list:
                if (not deduped or
                        _time_to_secs(c["hora"]) - _time_to_secs(deduped[-1]["hora"]) >= 180):
                    deduped.append(c)
            checadas_por_dia[dia] = deduped

        # ── Fila base+0: encabezados de día ──────────────────────────────────────
        ws.row_dimensions[base].height = 30
        for col_i, dia in enumerate(dias, start=1):
            cell = ws.cell(row=base, column=col_i)
            cell.value     = f"{DAY_ABBR[dia.weekday()]}\n{dia.day}"
            cell.font      = FONT_RED if dia.weekday() == 6 else FONT_BOLD
            cell.alignment = ALIGN_CENTER
            cell.border    = BORDER_THIN

        for col_i, label in ((15, "INCIDENCIAS"), (16, "OBSERVACIONES"), (17, "FIRMA DE CONFORMIDAD")):
            cell = ws.cell(row=base, column=col_i)
            cell.value     = label
            cell.font      = FONT_BOLD
            cell.alignment = ALIGN_CENTER
            cell.border    = BORDER_THIN

        # ── Fila base+1: ID y nombre ──────────────────────────────────────────────
        ws.cell(row=base+1, column=1).value  = "ID:"
        ws.cell(row=base+1, column=1).font   = FONT_BOLD
        ws.cell(row=base+1, column=2).value  = str(chec_id) if chec_id else ""
        ws.cell(row=base+1, column=4).value  = "Nombre:"
        ws.cell(row=base+1, column=4).font   = FONT_BOLD
        ws.cell(row=base+1, column=5).value  = trab["nombre"]
        _apply_borders_row(ws, base+1)

        # ── Filas base+2..base+5: datos de tiempo ─────────────────────────────────
        incidencias_count = 0
        obs_set: set[str] = set()

        for col_i, dia in enumerate(dias, start=1):
            checadas_hoy = checadas_por_dia.get(dia, [])
            horario_hoy  = horario_por_dia.get(dia)

            if not horario_hoy:
                # No es día laboral — celdas vacías con borde
                for row_off in range(2, 6):
                    ws.cell(row=base+row_off, column=col_i).border = BORDER_THIN
                continue

            hora_entrada = _td_to_time(horario_hoy["hora_entrada"])
            hora_salida  = _td_to_time(horario_hoy["hora_salida"])
            tiene_comida = bool(horario_hoy.get("tiene_comida", False))

            if not checadas_hoy:
                # Falta
                incidencias_count += 1
                cell = ws.cell(row=base+2, column=col_i)
                cell.value     = "F"
                cell.alignment = ALIGN_CENTER
                cell.border    = BORDER_THIN
                for row_off in (3, 4, 5):
                    ws.cell(row=base+row_off, column=col_i).border = BORDER_THIN
                continue

            r = _clasificar_checadas(checadas_hoy, hora_entrada, hora_salida, tiene_comida)

            if r["comida_incompleta"]:
                obs_set.add("OMISIÓN REGISTRO ALIMENTOS")

            # base+2: entrada
            cell_ent = ws.cell(row=base+2, column=col_i)
            cell_ent.value     = _fmt_hms(r["entrada"]) if r["entrada"] else None
            cell_ent.alignment = ALIGN_CENTER
            cell_ent.border    = BORDER_THIN
            if r["es_retardo"]:
                cell_ent.fill = FILL_PINK

            # base+3: salida a comer
            cell_csal = ws.cell(row=base+3, column=col_i)
            cell_csal.border = BORDER_THIN
            if tiene_comida and r["comida_sal"]:
                cell_csal.value     = _fmt_hms(r["comida_sal"])
                cell_csal.alignment = ALIGN_CENTER
                cell_csal.fill      = FILL_GREEN

            # base+4: regreso de comer
            cell_cent = ws.cell(row=base+4, column=col_i)
            cell_cent.border = BORDER_THIN
            if tiene_comida and r["comida_ent"]:
                cell_cent.value     = _fmt_hms(r["comida_ent"])
                cell_cent.alignment = ALIGN_CENTER
                cell_cent.fill      = FILL_GREEN

            # base+5: salida
            cell_sal = ws.cell(row=base+5, column=col_i)
            cell_sal.value     = _fmt_hms(r["salida"]) if r["salida"] else None
            cell_sal.alignment = ALIGN_CENTER
            cell_sal.border    = BORDER_THIN

        # ── Columnas O/P/Q mergeadas sobre base+2:base+5 ─────────────────────────
        # O — INCIDENCIAS
        ws.merge_cells(
            start_row=base+2, start_column=15, end_row=base+5, end_column=15
        )
        cell_o = ws.cell(row=base+2, column=15)
        cell_o.value     = incidencias_count if incidencias_count > 0 else None
        cell_o.alignment = ALIGN_CENTER
        cell_o.border    = BORDER_THIN

        # P — OBSERVACIONES
        ws.merge_cells(
            start_row=base+2, start_column=16, end_row=base+5, end_column=16
        )
        cell_p = ws.cell(row=base+2, column=16)
        cell_p.value     = "; ".join(sorted(obs_set)) if obs_set else None
        cell_p.alignment = ALIGN_WRAP
        cell_p.border    = BORDER_THIN

        # Q — FIRMA DE CONFORMIDAD
        ws.merge_cells(
            start_row=base+2, start_column=17, end_row=base+5, end_column=17
        )
        cell_q = ws.cell(row=base+2, column=17)
        cell_q.value  = None
        cell_q.border = BORDER_THIN

    # ── Orientación y escala ─────────────────────────────────────────────────────
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.scale       = 53

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
