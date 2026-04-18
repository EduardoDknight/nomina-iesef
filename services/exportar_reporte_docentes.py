"""
services/exportar_reporte_docentes.py
Reporte de Asistencia Docentes v2.0 — NEXO IESEF

Diseño mejorado frente al reporte v1 de administrativos:
  - Paleta de colores semántica por programa (misma que nomina_resumen)
  - Encabezado azul marino moderno
  - Fila ESTADO con íconos ✓ / F / ~ / V / ·
  - Columnas de resumen al final: BLOQUES ASIG. | ASISTENCIAS | FALTAS/PARC. | FIRMA

Estructura por docente (bloque de 7 filas):
  base+0  : Encabezados de día (L/Ma/Mi… + número) — azul profundo
  base+1  : ID  |  Nombre del docente  |  costo/hr   [resumen mergeado base+1..base+4]
  base+2  : ENTRADA registrada (HH:MM o "—" si falta, vacío si no hay clase)
  base+3  : SALIDA  registrada
  base+4  : ESTADO  (✓ asistencia | F falta | ~ parcial | V virtual | · sin clase)
  base+5/6: separador visual blanco (alto=4 px)

Clasificación de asistencia (por día con clase programada):
  - Primer timestamp = entrada, último = salida (si hay ≥2 checadas)
  - asistencia  : entrada dentro de ±10 min del inicio  Y  salida dentro de tolerancia final
  - parcial (~) : tiene entrada pero no salida válida (o viceversa)
  - falta   (F) : sin checadas ese día
  - virtual (V) : asignación de modalidad virtual (no usa checador)
  - sin_clase   : no hay horario programado para ese día
"""

import io
import logging
from datetime import date, timedelta
from collections import defaultdict
from typing import Dict, List, Tuple, Optional

from psycopg2.extras import RealDictCursor

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

logger = logging.getLogger(__name__)

# ── Paleta v2 ─────────────────────────────────────────────────────────────────
HDR_BG         = "1F3864"   # encabezado global: azul marino
HDR_FG         = "FFFFFF"
DAY_HDR_BG     = "2B4EAC"   # encabezado de columnas de día: azul profundo
DAY_HDR_FG     = "FFFFFF"

ESTADO_OK_BG   = "D0E8D0";  ESTADO_OK_FG   = "1B5E20"   # ✓ asistencia
ESTADO_F_BG    = "F8D7DA";  ESTADO_F_FG    = "9B1C1C"   # F falta
ESTADO_PARC_BG = "FFF3CD";  ESTADO_PARC_FG = "7C4E00"   # ~ parcial
ESTADO_VIR_BG  = "D1E7FF";  ESTADO_VIR_FG  = "0C3065"   # V virtual
ESTADO_SIN_BG  = "F3F4F6";  ESTADO_SIN_FG  = "9CA3AF"   # · sin clase
ENT_OK_BG      = "F0FDF4"   # celda de entrada cuando hay hora
ENT_FALTA_BG   = "FECACA"   # celda de entrada cuando falta / sin hora
SIN_CLASE_BG   = "F9FAFB"   # celda en días sin clase programada
RESUMEN_BG     = "EBF5FB"   # bloque de resumen
LABEL_BG       = "E8EEF8"   # etiquetas ENTRADA/SALIDA/ESTADO

FONT_NAME = "Arial"
_THIN = Side(style="thin")
_MED  = Side(style="medium")
BORDER_THIN = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Abreviaturas de día (weekday 0=lun … 6=dom)
DAY_ABBR = {0: "L", 1: "Ma", 2: "Mi", 3: "J", 4: "V", 5: "S", 6: "D"}
DIA_SEMANA = {
    0: "lunes", 1: "martes", 2: "miercoles",
    3: "jueves", 4: "viernes", 5: "sabado", 6: "domingo",
}
TOLERANCIA_ENTRADA_SEG = 10 * 60   # 10 min
TOLERANCIA_SALIDA_SEG  = 20 * 60   # tope 20 min (min(horas*10, 20))

# ── Paleta por programa (idéntica a exportar_nomina_resumen) ──────────────────
_COLORES_PROGRAMA: List[Tuple] = [
    (["BACHILLERATO", "PREPARATORIA"],
     "9DC3E6", "1F3864", "C9DDEF", "E4EFF8"),
    (["ENFERMERÍA", "ENFERMERIA", "ENFER"],
     "4472C4", "FFFFFF", "9DC3E6", "C9DDEF"),
    (["LENA", "NIVELACIÓN", "NIVELACION"],
     "8E72C4", "FFFFFF", "C5B8E0", "DDD6EF"),
    (["ESPECIALIDAD", "EEQX", "EECI", "EEPI", "EEGE", "ADSE"],
     "F4B183", "7F3000", "FAD4B0", "FDEBD8"),
    (["NUTRICIÓN", "NUTRICION", "NUTR"],
     "548235", "FFFFFF", "A9D18E", "D9EAD3"),
    (["MAESTRÍA", "MAESTRIA", "MSP", "MDIE", "MGDIS"],
     "2F5496", "FFFFFF", "8EA9C8", "C5D5E8"),
    (["CAMPO", "CLÍNICO", "CLINICO"],
     "C9302C", "FFFFFF", "F4CCCA", "FAE5E4"),
]
_FALLBACK = ("808080", "FFFFFF", "CCCCCC", "E8E8E8")

_ORDEN_PROG = [
    "BACHILLERATO", "PREPARATORIA",
    "ENFERMERÍA", "ENFERMERIA", "ENFER",
    "LENA", "NIVELACI",
    "ESPECIALIDAD", "EEQX", "EECI", "EEPI", "EEGE", "ADSE",
    "NUTRICI", "NUTR",
    "MAESTR", "MSP", "MDIE", "MGDIS",
    "CAMPO", "CLINICO",
]


def _color_programa(nombre: str) -> Tuple[str, str, str, str]:
    n = nombre.upper()
    for keywords, *colors in _COLORES_PROGRAMA:
        if any(k in n for k in keywords):
            return tuple(colors)
    return _FALLBACK


def _peso_prog(nombre: str) -> int:
    nu = nombre.upper()
    for i, pat in enumerate(_ORDEN_PROG):
        if pat in nu:
            return i
    return 999


# ── Helpers openpyxl ──────────────────────────────────────────────────────────

def _font(bold=False, color="000000", size=10) -> "Font":
    return Font(name=FONT_NAME, bold=bold, color=color, size=size)


def _fill(hex_color: str) -> "PatternFill":
    return PatternFill("solid", fgColor=hex_color)


def _put(ws, row: int, col: int, value=None,
         bg: str = "FFFFFF", fg: str = "000000",
         bold: bool = False, size: int = 10,
         halign: str = "center", valign: str = "center",
         wrap: bool = False) -> None:
    """Escribe una celda con estilo completo."""
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = _fill(bg)
    c.font      = _font(bold=bold, color=fg, size=size)
    c.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    c.border    = BORDER_THIN


# ── SQL ───────────────────────────────────────────────────────────────────────

SQL_QUINCENA = """
SELECT id, fecha_inicio::date, fecha_fin::date, ciclo, razon_social
FROM quincenas
WHERE id = %s
"""

# Un registro por (docente, programa) — DISTINCT ON para evitar duplicados
# cuando el docente tiene varias materias en el mismo programa
SQL_DOCENTES = """
SELECT DISTINCT ON (d.id, p.id)
    d.id                 AS docente_id,
    d.chec_id,
    d.nombre_completo,
    d.costo_hora_centro,
    d.costo_hora_instituto,
    p.id                 AS programa_id,
    p.nombre             AS programa_nombre,
    p.razon_social::text AS razon_social,
    a.modalidad::text    AS modalidad
FROM docentes d
JOIN asignaciones a ON a.docente_id = d.id
    AND a.activa = TRUE
JOIN materias m     ON m.id = a.materia_id
JOIN programas p    ON p.id = m.programa_id
WHERE d.activo = TRUE
ORDER BY d.id, p.id, a.modalidad
"""

# Todos los bloques horarios activos de los docentes dados
SQL_HORARIOS = """
SELECT
    a.docente_id,
    m.programa_id,
    a.modalidad::text    AS modalidad,
    hc.dia_semana::text  AS dia_semana,
    hc.hora_inicio,
    hc.hora_fin,
    hc.horas_bloque
FROM asignaciones a
JOIN materias m        ON m.id = a.materia_id
JOIN horario_clases hc ON hc.asignacion_id = a.id
WHERE a.docente_id = ANY(%s)
  AND a.activa = TRUE
ORDER BY a.docente_id, m.programa_id, hc.dia_semana, hc.hora_inicio
"""

SQL_CHECADAS = """
SELECT
    user_id,
    timestamp_checada::date AS dia,
    timestamp_checada::time AS hora,
    tipo_punch
FROM asistencias_checadas
WHERE user_id = ANY(%s)
  AND timestamp_checada::date >= %s
  AND timestamp_checada::date <= %s
ORDER BY user_id, timestamp_checada
"""


# ── Clasificación de asistencia ───────────────────────────────────────────────

def _td_to_time(t):
    from datetime import timedelta, time as dtime
    if isinstance(t, timedelta):
        total = int(t.total_seconds())
        h, rem = divmod(total, 3600)
        m, s   = divmod(rem, 60)
        return dtime(h % 24, m, s)
    return t


def _secs(t) -> int:
    from datetime import timedelta
    if isinstance(t, timedelta):
        return int(t.total_seconds())
    return t.hour * 3600 + t.minute * 60 + t.second


def _fmt(t) -> str:
    tt = _td_to_time(t)
    return f"{tt.hour:02d}:{tt.minute:02d}"


def _clasificar_dia(checadas: list, horario: dict) -> dict:
    """
    Clasifica las checadas de un día frente a un bloque horario.
    Retorna: {entrada_str, salida_str, estado}
    """
    if not checadas:
        return {"entrada_str": None, "salida_str": None, "estado": "falta"}

    sorted_c = sorted(checadas, key=lambda x: _secs(x["hora"]))
    entrada  = _td_to_time(sorted_c[0]["hora"])
    salida   = _td_to_time(sorted_c[-1]["hora"]) if len(sorted_c) >= 2 else None

    hi = _td_to_time(horario["hora_inicio"])
    hf = _td_to_time(horario["hora_fin"])
    hb = int(horario.get("horas_bloque") or 1)

    tol_sal = min(hb * 10, 20) * 60   # tope 20 min, en segundos

    ent_ok = _secs(entrada) <= _secs(hi) + TOLERANCIA_ENTRADA_SEG
    sal_ok = salida is not None and _secs(salida) >= _secs(hf) - tol_sal

    estado = "asistencia" if (ent_ok and sal_ok) else "parcial"
    return {
        "entrada_str": _fmt(entrada),
        "salida_str":  _fmt(salida) if salida else None,
        "estado":      estado,
    }


# ── Función pública ───────────────────────────────────────────────────────────

def generar_reporte_asistencia_docentes(conn, quincena_id: int) -> bytes:
    """
    Genera el Excel de Reporte de Asistencia Docentes v2.0.
    Devuelve los bytes del .xlsx.
    """
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl no está instalado")

    cur = conn.cursor(cursor_factory=RealDictCursor)

    # 1. Quincena ───────────────────────────────────────────────────────────────
    cur.execute(SQL_QUINCENA, (quincena_id,))
    q = cur.fetchone()
    if not q:
        raise ValueError(f"Quincena {quincena_id} no encontrada")

    fi: date  = q["fecha_inicio"]
    ff: date  = q["fecha_fin"]
    rs: str   = q["razon_social"]       # centro | instituto | ambas
    ciclo     = q["ciclo"] or ""

    dias: List[date] = []
    d = fi
    while d <= ff:
        dias.append(d)
        d += timedelta(days=1)
    n_dias = len(dias)

    # 2. Docentes + programas ──────────────────────────────────────────────────
    cur.execute(SQL_DOCENTES)
    all_rows = [dict(r) for r in cur.fetchall()]

    if rs != "ambas":
        all_rows = [r for r in all_rows if r["razon_social"] == rs]

    if not all_rows:
        raise ValueError("No hay docentes activos con asignaciones vigentes para esta quincena.")

    # prog_docentes[prog_nombre] = lista de dicts de docente (uno por docente)
    prog_docentes: Dict[str, List[dict]] = {}
    seen: set = set()

    for r in all_rows:
        pnom = r["programa_nombre"]
        key  = (r["docente_id"], pnom)
        if key in seen:
            continue
        seen.add(key)
        if pnom not in prog_docentes:
            prog_docentes[pnom] = []
        costo = (r["costo_hora_centro"]
                 if r["razon_social"] == "centro"
                 else r["costo_hora_instituto"])
        prog_docentes[pnom].append({
            "docente_id":  r["docente_id"],
            "chec_id":     r["chec_id"],
            "nombre":      r["nombre_completo"],
            "costo_hora":  costo,
            "modalidad":   r["modalidad"],
            "programa_id": r["programa_id"],
        })

    for pnom in prog_docentes:
        prog_docentes[pnom].sort(key=lambda x: x["nombre"])

    # 3. Horarios ──────────────────────────────────────────────────────────────
    doc_ids = list({r["docente_id"] for r in all_rows})
    cur.execute(SQL_HORARIOS, (doc_ids,))

    # horarios_map[docente_id][dia_semana_str][programa_id] = lista de bloques
    horarios_map: Dict = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    for h in cur.fetchall():
        horarios_map[h["docente_id"]][h["dia_semana"]][h["programa_id"]].append(dict(h))

    # 4. Checadas ──────────────────────────────────────────────────────────────
    chec_ids    = [r["chec_id"] for r in all_rows if r.get("chec_id") is not None]
    chec_to_doc = {r["chec_id"]: r["docente_id"]
                   for r in all_rows if r.get("chec_id") is not None}

    # checadas_map[docente_id][fecha] = lista de {hora, tipo_punch}
    checadas_map: Dict = defaultdict(lambda: defaultdict(list))
    if chec_ids:
        cur.execute(SQL_CHECADAS, (chec_ids, fi, ff))
        for c in cur.fetchall():
            did = chec_to_doc.get(c["user_id"])
            if did:
                checadas_map[did][c["dia"]].append({
                    "hora":       c["hora"],
                    "tipo_punch": c["tipo_punch"],
                })

    cur.close()

    # 5. Construir Excel ───────────────────────────────────────────────────────
    COL_BLOQUES = n_dias + 2
    COL_ASIST   = n_dias + 3
    COL_FALTAS  = n_dias + 4
    COL_FIRMA   = n_dias + 5
    NCOLS       = n_dias + 5

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ASISTENCIA"

    # Anchos de columna
    ws.column_dimensions["A"].width = 10.0
    for ci in range(2, n_dias + 2):
        ws.column_dimensions[get_column_letter(ci)].width = 9.0
    ws.column_dimensions[get_column_letter(COL_BLOQUES)].width = 10.0
    ws.column_dimensions[get_column_letter(COL_ASIST)].width   = 13.0
    ws.column_dimensions[get_column_letter(COL_FALTAS)].width  = 14.0
    ws.column_dimensions[get_column_letter(COL_FIRMA)].width   = 22.0

    # ── Títulos globales ──────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS)
    c = ws.cell(1, 1)
    c.value     = "REPORTE DE ASISTENCIA DOCENTES  —  NEXO IESEF  v2.0"
    c.fill      = _fill(HDR_BG)
    c.font      = Font(name=FONT_NAME, size=13, bold=True, color=HDR_FG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    rs_label = {"centro": "CENTRO", "instituto": "INSTITUTO",
                "ambas": "CENTRO + INSTITUTO"}.get(rs, rs.upper())
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NCOLS)
    c = ws.cell(2, 1)
    c.value     = (f"Quincena: {fi.strftime('%d/%m/%Y')} — {ff.strftime('%d/%m/%Y')}"
                   f"  ·  {rs_label}  ·  Ciclo: {ciclo}")
    c.fill      = _fill(DAY_HDR_BG)
    c.font      = Font(name=FONT_NAME, size=10, color=HDR_FG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 5   # spacer

    fila = 4

    # ── Programas ─────────────────────────────────────────────────────────────
    for pnom in sorted(prog_docentes.keys(), key=_peso_prog):
        docentes      = prog_docentes[pnom]
        sep_bg, sep_fg, bg_par, bg_impar = _color_programa(pnom)

        # Separador de programa
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=NCOLS)
        c = ws.cell(fila, 1)
        c.value     = f"  {pnom.upper()}"
        c.fill      = _fill(sep_bg)
        c.font      = Font(name=FONT_NAME, size=11, bold=True, color=sep_fg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = Border(bottom=Side(style="medium", color="FFFFFF"))
        ws.row_dimensions[fila].height = 20
        fila += 1

        # ── Bloque por docente (7 filas) ──────────────────────────────────────
        for doc_idx, doc in enumerate(docentes):
            did       = doc["docente_id"]
            chec_id   = doc["chec_id"]
            nombre    = doc["nombre"]
            costo     = doc["costo_hora"]
            modalidad = doc.get("modalidad", "presencial")
            prog_id   = doc["programa_id"]
            bg_info   = bg_par if doc_idx % 2 == 0 else bg_impar
            fg_info   = sep_fg if sep_fg != "FFFFFF" else "1F3864"

            # Pre-calcular datos de cada día
            dia_data: List[dict] = []
            bloques_asig = 0
            cnt_asist    = 0
            cnt_faltas   = 0

            for dia in dias:
                dname       = DIA_SEMANA[dia.weekday()]
                hors_hoy    = horarios_map[did][dname].get(prog_id, [])

                if not hors_hoy:
                    dia_data.append({"tiene_clase": False, "estado": "sin_clase",
                                     "entrada_str": None, "salida_str": None})
                    continue

                bloques_asig += 1
                horario = hors_hoy[0]   # primer bloque del día

                if modalidad == "virtual":
                    dia_data.append({"tiene_clase": True, "estado": "virtual",
                                     "entrada_str": None, "salida_str": None})
                    cnt_asist += 1
                    continue

                checadas_hoy = checadas_map[did].get(dia, [])
                res = _clasificar_dia(checadas_hoy, horario)
                if res["estado"] == "asistencia":
                    cnt_asist += 1
                else:
                    cnt_faltas += 1
                dia_data.append({"tiene_clase": True, **res})

            base = fila

            # base+0: Encabezados de días ────────────────────────────────────
            ws.row_dimensions[base].height = 28
            _put(ws, base, 1, None, bg=DAY_HDR_BG, fg=DAY_HDR_FG, bold=True)
            for ci, dia in enumerate(dias, 2):
                _put(ws, base, ci,
                     f"{DAY_ABBR[dia.weekday()]}\n{dia.day}",
                     bg=DAY_HDR_BG, fg=DAY_HDR_FG, bold=True, wrap=True)
            _put(ws, base, COL_BLOQUES, "BLOQUES\nASIG.",
                 bg=DAY_HDR_BG, fg=DAY_HDR_FG, bold=True, size=9, wrap=True)
            _put(ws, base, COL_ASIST,   "ASISTENCIAS",
                 bg=DAY_HDR_BG, fg=DAY_HDR_FG, bold=True, size=9)
            _put(ws, base, COL_FALTAS,  "FALTAS / PARC.",
                 bg=DAY_HDR_BG, fg=DAY_HDR_FG, bold=True, size=9)
            _put(ws, base, COL_FIRMA,   "FIRMA",
                 bg=DAY_HDR_BG, fg=DAY_HDR_FG, bold=True, size=9)

            # base+1: Info del docente + resumen (mergeado base+1..base+4) ────
            ws.row_dimensions[base + 1].height = 16
            _put(ws, base+1, 1, f"ID: {chec_id or '—'}",
                 bg=bg_info, fg=fg_info, bold=True, size=9)
            # Nombre y costo mergeado sobre columnas de días
            ws.merge_cells(start_row=base+1, start_column=2,
                           end_row=base+1, end_column=n_dias + 1)
            c = ws.cell(base+1, 2)
            costo_str = f"  ·  ${costo:,.0f}/hr" if costo else ""
            c.value     = f"  {nombre.upper()}{costo_str}"
            c.fill      = _fill(bg_info)
            c.font      = Font(name=FONT_NAME, size=10, bold=True, color=fg_info)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border    = BORDER_THIN
            # Resumen mergeado verticalmente base+1..base+4
            for col_r in (COL_BLOQUES, COL_ASIST, COL_FALTAS, COL_FIRMA):
                ws.merge_cells(start_row=base+1, start_column=col_r,
                               end_row=base+4, end_column=col_r)
            _put(ws, base+1, COL_BLOQUES, bloques_asig,
                 bg=RESUMEN_BG, fg="1F3864", bold=True, size=16)
            _put(ws, base+1, COL_ASIST, cnt_asist if cnt_asist > 0 else None,
                 bg=ESTADO_OK_BG, fg=ESTADO_OK_FG, bold=True, size=16)
            _put(ws, base+1, COL_FALTAS, cnt_faltas if cnt_faltas > 0 else None,
                 bg=ESTADO_F_BG if cnt_faltas > 0 else "FAFAFA",
                 fg=ESTADO_F_FG if cnt_faltas > 0 else ESTADO_SIN_FG,
                 bold=cnt_faltas > 0, size=16)
            _put(ws, base+1, COL_FIRMA, None, bg="FFFFFF")

            # base+2: ENTRADA ─────────────────────────────────────────────────
            ws.row_dimensions[base + 2].height = 15
            _put(ws, base+2, 1, "ENTRADA", bg=LABEL_BG, fg=DAY_HDR_BG,
                 bold=True, size=8, halign="right")
            for ci, dd in enumerate(dia_data, 2):
                if not dd["tiene_clase"]:
                    _put(ws, base+2, ci, None, bg=SIN_CLASE_BG)
                elif dd["estado"] == "virtual":
                    _put(ws, base+2, ci, "VIRTUAL",
                         bg=ESTADO_VIR_BG, fg=ESTADO_VIR_FG, size=7)
                elif dd["estado"] == "falta":
                    _put(ws, base+2, ci, "—",
                         bg=ENT_FALTA_BG, fg=ESTADO_F_FG, bold=True)
                else:
                    has_ent = bool(dd["entrada_str"])
                    _put(ws, base+2, ci, dd["entrada_str"] or "—",
                         bg=ENT_OK_BG if has_ent else ENT_FALTA_BG,
                         fg=ESTADO_OK_FG if has_ent else ESTADO_F_FG,
                         size=9)

            # base+3: SALIDA ──────────────────────────────────────────────────
            ws.row_dimensions[base + 3].height = 15
            _put(ws, base+3, 1, "SALIDA", bg=LABEL_BG, fg=DAY_HDR_BG,
                 bold=True, size=8, halign="right")
            for ci, dd in enumerate(dia_data, 2):
                if not dd["tiene_clase"]:
                    _put(ws, base+3, ci, None, bg=SIN_CLASE_BG)
                elif dd["estado"] == "virtual":
                    _put(ws, base+3, ci, None, bg=ESTADO_VIR_BG)
                elif dd["estado"] == "falta":
                    _put(ws, base+3, ci, None, bg=ENT_FALTA_BG)
                else:
                    has_sal = bool(dd["salida_str"])
                    _put(ws, base+3, ci, dd["salida_str"] or "—",
                         bg=ENT_OK_BG if has_sal else ESTADO_PARC_BG,
                         fg=ESTADO_OK_FG if has_sal else ESTADO_PARC_FG,
                         size=9)

            # base+4: ESTADO ──────────────────────────────────────────────────
            ws.row_dimensions[base + 4].height = 16
            _put(ws, base+4, 1, "ESTADO", bg=LABEL_BG, fg=DAY_HDR_BG,
                 bold=True, size=8, halign="right")
            ESTADOS = {
                "asistencia": ("✓", ESTADO_OK_BG,   ESTADO_OK_FG,   True),
                "falta":      ("F",  ESTADO_F_BG,    ESTADO_F_FG,    True),
                "parcial":    ("~",  ESTADO_PARC_BG, ESTADO_PARC_FG, True),
                "virtual":    ("V",  ESTADO_VIR_BG,  ESTADO_VIR_FG,  True),
                "sin_clase":  ("·",  ESTADO_SIN_BG,  ESTADO_SIN_FG,  False),
            }
            for ci, dd in enumerate(dia_data, 2):
                sym, bg_e, fg_e, bl = ESTADOS.get(
                    dd["estado"], ("·", ESTADO_SIN_BG, ESTADO_SIN_FG, False))
                _put(ws, base+4, ci, sym, bg=bg_e, fg=fg_e, bold=bl, size=11)

            # base+5, base+6: separador visual (alto=4 px, blanco) ────────────
            for ro in (5, 6):
                ws.row_dimensions[base + ro].height = 4
                for ci in range(1, NCOLS + 1):
                    ws.cell(base + ro, ci).fill = _fill("FFFFFF")

            fila += 7

    # ── Configuración de página ────────────────────────────────────────────────
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.freeze_panes = "B4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
