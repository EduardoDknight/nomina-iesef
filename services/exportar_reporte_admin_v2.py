"""
services/exportar_reporte_admin_v2.py
Reporte de Checador Administrativos v2.0 — NEXO IESEF

Diseño minimalista de marca:
  · Encabezados dark-navy con texto blanco limpio
  · Nombre del trabajador prominente (fila dedicada, fuente grande)
  · Domingo atenuado en gris — día no laborable
  · Columnas de resumen: RETARDOS | FALTAS (con desglose N+M) | INCIDENCIAS | OBSERVACIONES | FIRMA
  · "OMISIÓN REGISTRO ALIMENTOS" automático cuando comida_incompleta
  · Algoritmo por zonas: comida clasificada primero (±90 min del punto medio);
    entrada/salida del resto — MB360 envía casi todo como tipo_punch=0
  · Impresión horizontal carta (landscape, fit-to-width, márgenes ajustados)
"""

from io import BytesIO
from datetime import date, time, timedelta
from collections import defaultdict
from typing import Optional

import psycopg2
from psycopg2.extras import RealDictCursor
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from config import settings

# ── Paleta minimalista ────────────────────────────────────────────────────────
HDR_BG        = "0F172A"   # casi negro — barra de título
HDR_FG        = "FFFFFF"
SUB_HDR_BG    = "1E293B"   # dark slate — subtítulo / fechas
SUB_HDR_FG    = "94A3B8"   # texto apagado

DAY_HDR_BG    = "1E3A5F"   # deep navy — encabezados de días (lunes–sábado)
DAY_HDR_FG    = "FFFFFF"
SUN_HDR_BG    = "64748B"   # slate-500 — domingo (no laborable)
SUN_HDR_FG    = "E2E8F0"
SUN_CELL_BG   = "E2E8F0"   # celdas de domingo

NOMBRE_BG     = "1E3A5F"   # mismo que DAY_HDR — fila de nombre
NOMBRE_FG     = "FFFFFF"
NOMBRE_ID_FG  = "93C5FD"   # blue-300 — ID pequeño

ENT_OK_BG     = "F8FAFC"   # casi blanco — entrada normal
ENT_TARD_BG   = "FEF3C7"   # amber-100 — retardo
ENT_TARD_FG   = "92400E"   # amber-800
FALTA_BG      = "FEE2E2"   # red-100 — falta
FALTA_FG      = "B91C1C"   # red-700
COMIDA_BG     = "DCFCE7"   # green-100 — comida
COMIDA_FG     = "166534"   # green-800
SAL_BG        = "F8FAFC"   # casi blanco — salida
SIN_HORARIO   = "F1F5F9"   # slate-100 — día sin horario definido

RETARDOS_BG   = "FEF3C7"
RETARDOS_FG   = "92400E"
FALTAS_BG     = "FEE2E2"
FALTAS_FG     = "B91C1C"
OBS_BG        = "FFFBEB"   # warm-50 — fondo observaciones
OBS_FG        = "B45309"   # amber-700
FIRMA_BG      = "FFFFFF"
LABEL_BG      = "F1F5F9"   # slate-100 — etiqueta col A
LABEL_FG      = "334155"   # slate-700

FONT_NAME = "Arial"
_THIN = Side(style="thin")
BORDER_THIN = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

DAY_ABBR = {0: "L", 1: "M", 2: "M", 3: "J", 4: "V", 5: "S", 6: "D"}
DAY_COL  = {
    0: "lunes", 1: "martes", 2: "miercoles", 3: "jueves",
    4: "viernes", 5: "sabado", 6: "domingo",
}

TOLERANCIA_COMIDA_SEG = 90 * 60   # ±90 min alrededor del punto medio de la jornada


# ── Helpers ───────────────────────────────────────────────────────────────────

def _get_conn():
    return psycopg2.connect(settings.database_url_nomina, cursor_factory=RealDictCursor)


def _td_to_time(t) -> time:
    if isinstance(t, timedelta):
        total = int(t.total_seconds())
        h, rem = divmod(total, 3600)
        m, s   = divmod(rem, 60)
        return time(h % 24, m, s)
    return t


def _time_to_secs(t) -> int:
    if isinstance(t, timedelta):
        return int(t.total_seconds())
    return t.hour * 3600 + t.minute * 60 + t.second


def _fmt_hms(t) -> str:
    tt = _td_to_time(t)
    return f"{tt.hour:02d}:{tt.minute:02d}:{tt.second:02d}"


def _font(bold=False, color="000000", size=9) -> Font:
    return Font(name=FONT_NAME, bold=bold, color=color, size=size)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _put(ws, row: int, col: int, value=None,
         bg: str = "FFFFFF", fg: str = "000000",
         bold: bool = False, size: int = 9,
         halign: str = "center", valign: str = "center",
         wrap: bool = False, border: bool = True) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = _fill(bg)
    c.font      = _font(bold=bold, color=fg, size=size)
    c.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    if border:
        c.border = BORDER_THIN


# ── Clasificación por zonas (MB360 casi siempre tipo_punch=0) ─────────────────

def _clasificar_checadas(checadas_raw: list, hora_entrada: time,
                          hora_salida: time, tiene_comida: bool) -> dict:
    """
    Clasifica las checadas del día en entrada / comida_sal / comida_ent / salida.

    Estrategia:
      1. Si el dispositivo envió tipos explícitos de comida (2/3) → confiar en ellos.
      2. Si tiene_comida: clasificar por zonas.
         - Zona comida = checadas dentro de ±90 min del punto medio de la jornada.
         - Zona trabajo (fuera) → antes del mediodía = entrada; después = salida.
      3. Sin comida: primera=entrada, última=salida.
    """
    checadas_sorted = sorted(checadas_raw, key=lambda x: _time_to_secs(x["hora"]))
    all_horas = [_td_to_time(c["hora"]) for c in checadas_sorted]

    entrada_checada:    Optional[time] = None
    salida_checada:     Optional[time] = None
    comida_sal_checada: Optional[time] = None
    comida_ent_checada: Optional[time] = None

    by_type: dict = defaultdict(list)
    for c in checadas_sorted:
        by_type[c["tipo_punch"]].append(_td_to_time(c["hora"]))
    has_dedicated_comida = bool(by_type.get(2) or by_type.get(3))

    if has_dedicated_comida:
        comida_sal_checada = by_type[2][0]  if by_type.get(2) else None
        comida_ent_checada = by_type[3][0]  if by_type.get(3) else None
        entrada_checada    = by_type[0][0]  if by_type.get(0) else None
        salida_checada     = by_type[1][-1] if by_type.get(1) else None
    elif tiene_comida:
        # Tolerancia proporcional al turno: 25% de la duración, mín 30 min, máx 90 min
        shift_secs = abs(_time_to_secs(hora_salida) - _time_to_secs(hora_entrada))
        tol        = max(30 * 60, min(TOLERANCIA_COMIDA_SEG, shift_secs // 4))
        mid_secs   = (_time_to_secs(hora_entrada) + _time_to_secs(hora_salida)) // 2

        zona_comida  = [h for h in all_horas
                        if abs(_time_to_secs(h) - mid_secs) <= tol]
        zona_trabajo = [h for h in all_horas
                        if abs(_time_to_secs(h) - mid_secs) > tol]

        if zona_comida:
            comida_sal_checada = zona_comida[0]
        if len(zona_comida) >= 2:
            comida_ent_checada = zona_comida[1]

        antes   = [h for h in zona_trabajo if _time_to_secs(h) < mid_secs]
        despues = [h for h in zona_trabajo if _time_to_secs(h) > mid_secs]
        if antes:
            entrada_checada = antes[0]
        if despues:
            salida_checada = despues[-1]
    else:
        entrada_checada = all_horas[0]        if all_horas else None
        salida_checada  = all_horas[-1] if len(all_horas) >= 2 else None

    es_retardo = False
    if entrada_checada:
        es_retardo = _time_to_secs(entrada_checada) > _time_to_secs(hora_entrada) + 10 * 60

    comida_incompleta = (
        tiene_comida and
        (comida_sal_checada is None or comida_ent_checada is None)
    )
    return {
        "entrada":           entrada_checada,
        "salida":            salida_checada,
        "comida_sal":        comida_sal_checada,
        "comida_ent":        comida_ent_checada,
        "es_retardo":        es_retardo,
        "comida_incompleta": comida_incompleta,
    }


# ── Función principal ─────────────────────────────────────────────────────────

def generar_reporte_admin_v2(periodo_id: int) -> bytes:
    """Genera el Excel de Reporte Administrativos v2.0 — diseño minimalista."""
    conn = _get_conn()
    cur  = conn.cursor()

    # 1. Período
    cur.execute(
        "SELECT id, nombre, fecha_inicio, fecha_fin FROM periodos_admin WHERE id = %s",
        (periodo_id,)
    )
    periodo = cur.fetchone()
    if not periodo:
        conn.close()
        raise ValueError(f"Período {periodo_id} no encontrado")

    fecha_inicio: date = periodo["fecha_inicio"]
    fecha_fin:    date = periodo["fecha_fin"]
    nombre_periodo     = periodo["nombre"] or f"{fecha_inicio} – {fecha_fin}"

    dias: list[date] = []
    d = fecha_inicio
    while d <= fecha_fin:
        dias.append(d)
        d += timedelta(days=1)
    n_dias = len(dias)

    # 2. Trabajadores activos
    cur.execute(
        "SELECT id, chec_id, nombre, cargo FROM trabajadores WHERE activo = TRUE ORDER BY nombre"
    )
    trabajadores = [dict(r) for r in cur.fetchall()]

    # 3. Horarios
    trab_ids = [t["id"] for t in trabajadores]
    horarios_por_trab: dict = defaultdict(list)
    if trab_ids:
        cur.execute(
            """SELECT trabajador_id, lunes, martes, miercoles, jueves, viernes,
                      sabado, domingo, hora_entrada, hora_salida, tiene_comida
               FROM horarios_trabajador
               WHERE trabajador_id = ANY(%s)""",
            (trab_ids,)
        )
        for h in cur.fetchall():
            horarios_por_trab[h["trabajador_id"]].append(dict(h))

    # 4. Checadas del período
    chec_ids = [t["chec_id"] for t in trabajadores if t["chec_id"] is not None]
    checadas_por_chec_dia: dict = defaultdict(lambda: defaultdict(list))
    if chec_ids:
        cur.execute(
            """SELECT user_id,
                      timestamp_checada::date AS dia,
                      timestamp_checada::time AS hora,
                      tipo_punch
               FROM asistencias_checadas
               WHERE user_id = ANY(%s)
                 AND timestamp_checada::date >= %s
                 AND timestamp_checada::date <= %s
               ORDER BY timestamp_checada""",
            (chec_ids, fecha_inicio, fecha_fin)
        )
        for c in cur.fetchall():
            checadas_por_chec_dia[c["user_id"]][c["dia"]].append({
                "hora":       c["hora"],
                "tipo_punch": c["tipo_punch"],
            })

    conn.close()

    # 5. Construir Excel ───────────────────────────────────────────────────────
    #   Columnas: A(etiqueta) · días · RETARDOS · FALTAS · INCIDENCIAS · OBSERVACIONES · FIRMA
    COL_RETARDOS    = n_dias + 2
    COL_FALTAS      = n_dias + 3
    COL_INCIDENCIAS = n_dias + 4
    COL_OBS         = n_dias + 5
    COL_FIRMA       = n_dias + 6
    NCOLS           = n_dias + 6

    wb = Workbook()
    ws = wb.active
    ws.title = "REPORTE"

    # ── Anchos ───────────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 10.5     # etiqueta
    for ci in range(2, n_dias + 2):
        is_sun = dias[ci - 2].weekday() == 6
        ws.column_dimensions[get_column_letter(ci)].width = 6.0 if is_sun else 9.2
    ws.column_dimensions[get_column_letter(COL_RETARDOS)].width    = 9.0
    ws.column_dimensions[get_column_letter(COL_FALTAS)].width      = 11.0
    ws.column_dimensions[get_column_letter(COL_INCIDENCIAS)].width = 11.0
    ws.column_dimensions[get_column_letter(COL_OBS)].width         = 16.0
    ws.column_dimensions[get_column_letter(COL_FIRMA)].width       = 22.0

    # ── Fila 1: Título ────────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS)
    c = ws.cell(1, 1)
    c.value     = "REPORTE DE CHECADOR  ·  PERSONAL ADMINISTRATIVO"
    c.fill      = _fill(HDR_BG)
    c.font      = Font(name=FONT_NAME, size=12, bold=True, color=HDR_FG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # ── Fila 2: Período ───────────────────────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NCOLS)
    c = ws.cell(2, 1)
    c.value     = (f"{nombre_periodo}  ·  "
                   f"{fecha_inicio.strftime('%d/%m/%Y')} — {fecha_fin.strftime('%d/%m/%Y')}")
    c.fill      = _fill(SUB_HDR_BG)
    c.font      = Font(name=FONT_NAME, size=9, italic=True, color=SUB_HDR_FG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 14

    # ── Fila 3: espaciador ────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 4

    # ── Bloque por trabajador (8 filas) ───────────────────────────────────────
    #   base+0  encabezados de días
    #   base+1  ID · Nombre (prominente)
    #   base+2  ENTRADA
    #   base+3  SAL.COMER
    #   base+4  REG.COMER
    #   base+5  SALIDA
    #   base+6  separador oscuro delgado
    #   base+7  microespacio
    for idx, trab in enumerate(trabajadores):
        base    = 4 + idx * 8
        chec_id = trab.get("chec_id")

        # Horario efectivo por día de la semana
        horarios = horarios_por_trab.get(trab["id"], [])
        horario_por_dia: dict = {}
        for dia in dias:
            col_nombre = DAY_COL[dia.weekday()]
            for h in horarios:
                if h.get(col_nombre):
                    horario_por_dia[dia] = h
                    break

        # Checadas del trabajador (dedup 3 min)
        raw_por_dia = checadas_por_chec_dia.get(chec_id, {}) if chec_id else {}
        checadas_por_dia: dict = {}
        for dia, raw_list in raw_por_dia.items():
            sl = sorted(raw_list, key=lambda x: _time_to_secs(x["hora"]))
            deduped = []
            for cc in sl:
                if (not deduped or
                        _time_to_secs(cc["hora"]) - _time_to_secs(deduped[-1]["hora"]) >= 180):
                    deduped.append(cc)
            checadas_por_dia[dia] = deduped

        # ── base+0: Encabezados de días ───────────────────────────────────────
        ws.row_dimensions[base].height = 22
        _put(ws, base, 1, None, bg=DAY_HDR_BG)
        for col_i, dia in enumerate(dias, 2):
            is_sun = dia.weekday() == 6
            hbg = SUN_HDR_BG if is_sun else DAY_HDR_BG
            hfg = SUN_HDR_FG if is_sun else DAY_HDR_FG
            _put(ws, base, col_i,
                 f"{DAY_ABBR[dia.weekday()]}\n{dia.day}",
                 bg=hbg, fg=hfg,
                 bold=not is_sun, size=8, wrap=True)
        # Cabeceras de resumen en la misma fila base
        for col_r, lbl in (
            (COL_RETARDOS,    "RETARDOS"),
            (COL_FALTAS,      "FALTAS"),
            (COL_INCIDENCIAS, "INCIDENCIAS"),
            (COL_OBS,         "OBSERVACIONES"),
            (COL_FIRMA,       "FIRMA DE CONFORMIDAD"),
        ):
            _put(ws, base, col_r, lbl,
                 bg=DAY_HDR_BG, fg=DAY_HDR_FG, bold=True, size=8)

        # ── base+1: Nombre (prominente) ───────────────────────────────────────
        ws.row_dimensions[base + 1].height = 20
        _put(ws, base+1, 1,
             f"#{chec_id or '—'}",
             bg=NOMBRE_BG, fg=NOMBRE_ID_FG, bold=False, size=8)
        ws.merge_cells(start_row=base+1, start_column=2,
                       end_row=base+1, end_column=n_dias+1)
        c = ws.cell(base+1, 2)
        c.value     = trab["nombre"].upper()
        c.fill      = _fill(NOMBRE_BG)
        c.font      = Font(name=FONT_NAME, size=11, bold=True, color=NOMBRE_FG)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border    = BORDER_THIN

        # Merge vertical base+1..base+5 para columnas de resumen
        for col_r in (COL_RETARDOS, COL_FALTAS, COL_INCIDENCIAS, COL_OBS, COL_FIRMA):
            ws.merge_cells(start_row=base+1, start_column=col_r,
                           end_row=base+5, end_column=col_r)

        # ── base+2..base+5: etiquetas fijas col A ─────────────────────────────
        LABELS = {base+2: "ENTRADA", base+3: "SAL.COMER",
                  base+4: "REG.COMER", base+5: "SALIDA"}
        for row_off in (2, 3, 4, 5):
            ws.row_dimensions[base + row_off].height = 14
            _put(ws, base+row_off, 1, LABELS[base+row_off],
                 bg=LABEL_BG, fg=LABEL_FG, bold=False, size=8, halign="right")

        # ── Datos por día ─────────────────────────────────────────────────────
        cnt_retardos         = 0
        cnt_faltas           = 0
        comida_incompleta_ok = False   # algún día tuvo omisión registro alimentos

        for col_i, dia in enumerate(dias, 2):
            is_sun       = dia.weekday() == 6
            checadas_hoy = checadas_por_dia.get(dia, [])
            horario_hoy  = horario_por_dia.get(dia)

            # Domingo: celda gris, sin borde
            if is_sun:
                for row_off in (2, 3, 4, 5):
                    _put(ws, base+row_off, col_i, None,
                         bg=SUN_CELL_BG, border=False)
                continue

            # Día sin horario definido
            if not horario_hoy:
                for row_off in (2, 3, 4, 5):
                    _put(ws, base+row_off, col_i, None, bg=SIN_HORARIO)
                continue

            hora_ent_h   = _td_to_time(horario_hoy["hora_entrada"])
            hora_sal_h   = _td_to_time(horario_hoy["hora_salida"])
            tiene_comida = bool(horario_hoy.get("tiene_comida", False))

            # Falta completa (nadie checó)
            if not checadas_hoy:
                cnt_faltas += 1
                _put(ws, base+2, col_i, "FALTA",
                     bg=FALTA_BG, fg=FALTA_FG, bold=True, size=8)
                for row_off in (3, 4, 5):
                    _put(ws, base+row_off, col_i, None, bg=FALTA_BG)
                continue

            r = _clasificar_checadas(checadas_hoy, hora_ent_h, hora_sal_h, tiene_comida)
            if r["comida_incompleta"]:
                comida_incompleta_ok = True

            # ENTRADA
            if r["es_retardo"]:
                cnt_retardos += 1
                _put(ws, base+2, col_i,
                     _fmt_hms(r["entrada"]) if r["entrada"] else None,
                     bg=ENT_TARD_BG, fg=ENT_TARD_FG, bold=True, size=9)
            else:
                _put(ws, base+2, col_i,
                     _fmt_hms(r["entrada"]) if r["entrada"] else None,
                     bg=ENT_OK_BG, size=9)

            # SAL.COMER — verde solo si hay checada real
            if tiene_comida and r["comida_sal"]:
                _put(ws, base+3, col_i, _fmt_hms(r["comida_sal"]),
                     bg=COMIDA_BG, fg=COMIDA_FG, size=9)
            else:
                _put(ws, base+3, col_i, None, bg="FFFFFF")

            # REG.COMER — verde solo si hay checada real
            if tiene_comida and r["comida_ent"]:
                _put(ws, base+4, col_i, _fmt_hms(r["comida_ent"]),
                     bg=COMIDA_BG, fg=COMIDA_FG, size=9)
            else:
                _put(ws, base+4, col_i, None, bg="FFFFFF")

            # SALIDA
            _put(ws, base+5, col_i,
                 _fmt_hms(r["salida"]) if r["salida"] else None,
                 bg=SAL_BG, size=9)

        # ── Columnas de resumen (ya mergeadas) ────────────────────────────────
        # RETARDOS
        _put(ws, base+1, COL_RETARDOS,
             cnt_retardos if cnt_retardos else None,
             bg=RETARDOS_BG if cnt_retardos else "F8FAFC",
             fg=RETARDOS_FG if cnt_retardos else "94A3B8",
             bold=bool(cnt_retardos), size=9)

        # FALTAS — formato "total (directas+por_retardos)" si aplica
        faltas_x_retardos = cnt_retardos // 3
        total_faltas      = cnt_faltas + faltas_x_retardos
        if total_faltas:
            if faltas_x_retardos:
                falta_val = f"{total_faltas} ({cnt_faltas}+{faltas_x_retardos})"
            else:
                falta_val = str(total_faltas)
        else:
            falta_val = None
        _put(ws, base+1, COL_FALTAS,
             falta_val,
             bg="FFFFFF" if total_faltas else "F8FAFC",
             fg=FALTAS_FG if total_faltas else "94A3B8",
             bold=bool(total_faltas), size=9)

        # INCIDENCIAS (llenado manual)
        _put(ws, base+1, COL_INCIDENCIAS, None, bg="F8FAFC")

        # OBSERVACIONES — automático si hubo omisión comida
        obs_txt = "OMISIÓN REGISTRO\nALIMENTOS" if comida_incompleta_ok else None
        _put(ws, base+1, COL_OBS, obs_txt,
             bg="FFFFFF" if obs_txt else "F8FAFC",
             fg=OBS_FG if obs_txt else "94A3B8",
             bold=False, size=8, halign="center", valign="center", wrap=True)

        # FIRMA
        _put(ws, base+1, COL_FIRMA, None, bg=FIRMA_BG)

        # ── base+6: línea separadora oscura ──────────────────────────────────
        ws.row_dimensions[base + 6].height = 3
        for ci in range(1, NCOLS + 1):
            ws.cell(base + 6, ci).fill = _fill(HDR_BG)

        ws.row_dimensions[base + 7].height = 1

    # ── Configuración de página — horizontal carta ────────────────────────────
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize   = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(
        left=0.35, right=0.35, top=0.45, bottom=0.45,
        header=0.2, footer=0.2,
    )
    ws.freeze_panes = "B4"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
