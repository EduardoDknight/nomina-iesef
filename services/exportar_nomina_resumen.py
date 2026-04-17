"""
Resumen de Nómina — IESEF
Dos hojas: CENTRO e INSTITUTO (agrupado por programa).

Diseño:
  - Encabezado de columnas: azul marino, texto blanco bold
  - Cada programa tiene su propio color pastel (10 en ciclo)
  - Separador de programa: tono medio del color — nombre bold
  - Filas de datos: alternando tono claro / más claro del mismo color
  - Descuentos: texto rojo, fondo del color del programa (no fondo rojo)
  - Totales: gris, bold
  - Fuente Arial consistente
"""
import io
import logging
from decimal import Decimal, ROUND_HALF_UP
from typing import Dict, List, Optional, Tuple

from psycopg2.extras import RealDictCursor

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

logger = logging.getLogger(__name__)

# ── Paleta fija ───────────────────────────────────────────────────────────────
HDR_BG     = "1F3864"
HDR_FG     = "FFFFFF"
TOT_BG     = "EDEDED"
RED_TEXT   = "C00000"
BLACK_TEXT = "000000"
FONT_NAME  = "Arial"

# Colores por programa: (sep_bg, sep_fg, row_par, row_impar)
# sep_fg = color del texto en la fila separadora (blanco sobre oscuro, negro sobre claro)
_COLORES_PROGRAMA = [
    # keywords_upper           sep_bg    sep_fg    row_par   row_impar
    (["BACHILLERATO",
      "PREPARATORIA"],        "9DC3E6", "1F3864", "C9DDEF", "E4EFF8"),  # azul muy claro
    (["ENFERMERÍA",
      "ENFERMERIA",
      "ENFER"],               "4472C4", "FFFFFF", "9DC3E6", "C9DDEF"),  # azul medio
    (["LENA",
      "NIVELACIÓN",
      "NIVELACION"],          "8E72C4", "FFFFFF", "C5B8E0", "DDD6EF"),  # lavanda
    (["ESPECIALIDAD",
      "EEQX","EECI",
      "EEPI","EEGE","ADSE"],  "F4B183", "7F3000", "FAD4B0", "FDEBD8"),  # naranja
    (["NUTRICIÓN",
      "NUTRICION",
      "NUTR"],                "548235", "FFFFFF", "A9D18E", "D9EAD3"),  # verde
    (["MAESTRÍA",
      "MAESTRIA",
      "MSP","MDIE","MGDIS"],  "2F5496", "FFFFFF", "8EA9C8", "C5D5E8"),  # azul marino
    (["CAMPO",
      "CLÍNICO",
      "CLINICO"],             "C9302C", "FFFFFF", "F4CCCA", "FAE5E4"),  # rojo suave
]
# Fallback para programas no reconocidos (cicla por índice)
_PALETA_FALLBACK: List[Tuple[str, str, str, str]] = [
    ("808080", "FFFFFF", "CCCCCC", "E8E8E8"),
    ("76BDBD", "1F3864", "AEDADA", "D6EDED"),
    ("C27BA0", "FFFFFF", "DDA8C8", "EED4E4"),
]


def _color_programa(nombre: str, fallback_idx: int = 0) -> Tuple[str, str, str, str]:
    """Retorna (sep_bg, sep_fg, row_par, row_impar) según nombre del programa."""
    n = nombre.upper()
    for keywords, sep_bg, sep_fg, row_par, row_impar in _COLORES_PROGRAMA:
        if any(k in n for k in keywords):
            return sep_bg, sep_fg, row_par, row_impar
    return _PALETA_FALLBACK[fallback_idx % len(_PALETA_FALLBACK)]

# Orden canónico programas INSTITUTO
ORDEN_INSTITUTO = [
    "ENFERMERÍA", "ENFERMERIA",
    "LICENCIATURA EN ENFERMERÍA", "LICENCIATURA EN ENFERMERIA",
    "LENA", "LIC. ENFERMERÍA NIVELACIÓN ACADÉMICA",
    "ESPECIALIDADES",
    "NUTRICIÓN", "NUTRICION", "LICENCIATURA EN NUTRICIÓN",
    "MAESTRÍAS", "MAESTRIAS",
    "CAMPO CLÍNICO", "CAMPO CLINICO",
]


# ── Estilos base ───────────────────────────────────────────────────────────────

def _font(bold=False, color=BLACK_TEXT, size=10) -> Font:
    return Font(name=FONT_NAME, bold=bold, color=color, size=size)

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _money_fmt() -> str:
    return '"$"#,##0.00;[Red]("$"#,##0.00);"-"'

def _hdr_cell(cell, texto: str) -> None:
    cell.value     = texto
    cell.fill      = _fill(HDR_BG)
    cell.font      = _font(bold=True, color=HDR_FG, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = Border(bottom=Side(style="medium", color="FFFFFF"))

def _sep_cell(ws, fila: int, n_cols: int, texto: str,
              bg_sep: str, fg_sep: str = "1F3864") -> None:
    """Separador de programa con el color del bloque."""
    for col in range(1, n_cols + 1):
        c = ws.cell(fila, col)
        c.fill   = _fill(bg_sep)
        c.border = Border(bottom=Side(style="thin", color="AAAAAA"))
    ws.cell(fila, 1).value     = texto.upper()
    ws.cell(fila, 1).font      = _font(bold=True, color=fg_sep, size=10)
    ws.cell(fila, 1).alignment = Alignment(horizontal="left", vertical="center")

def _data_cell(ws, fila: int, col: int, valor,
               bg: str = "FFFFFF", money: bool = False,
               rojo: bool = False) -> None:
    c = ws.cell(fila, col, valor)
    c.fill      = _fill(bg)
    c.font      = _font(color=RED_TEXT if rojo else BLACK_TEXT)
    c.alignment = Alignment(horizontal="right" if money else "left",
                            vertical="center")
    if money and valor is not None:
        c.number_format = _money_fmt()

def _tot_cell(ws, fila: int, col: int, valor,
              money: bool = False) -> None:
    c = ws.cell(fila, col, valor)
    c.fill      = _fill(TOT_BG)
    c.font      = _font(bold=True)
    c.alignment = Alignment(horizontal="right" if money else "left",
                            vertical="center")
    if money:
        c.number_format = _money_fmt()
    c.border = Border(top=Side(style="medium"), bottom=Side(style="medium"))

def _auto_ancho(ws, anchos_min: Dict[int, int]) -> None:
    for col_idx, min_w in anchos_min.items():
        letra   = get_column_letter(col_idx)
        max_len = min_w
        for row in ws.iter_rows():
            cell = row[col_idx - 1]
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letra].width = min(max_len + 2, 46)


# ── SQL ────────────────────────────────────────────────────────────────────────

SQL_QUINCENA = """
SELECT id, fecha_inicio, fecha_fin, ciclo, razon_social
FROM quincenas WHERE id = %s
"""

# SIN filtro por d.adscripcion — la clasificación centro/instituto se hace
# a partir de programas.razon_social en nomina_detalle_programa.
# Esto resuelve el bug donde docentes con adscripcion='instituto' pero con
# asignaciones en Bachillerato (centro) no aparecían en el Excel de Centro.
SQL_NOMINA = """
SELECT
    nq.docente_id,
    d.nombre_completo,
    d.noi,
    d.adscripcion,
    nq.honorarios,
    COALESCE(nq.horas_descuento * nq.costo_hora_promedio, 0) AS descuentos_total,
    nq.ajustes,
    nq.total_final
FROM nomina_quincena nq
JOIN docentes d ON d.id = nq.docente_id
WHERE nq.quincena_id = %s
ORDER BY d.nombre_completo
"""

# Datos REALES per-programa desde nomina_detalle_programa.
# Antes se usaba SQL_ASIGNACIONES con distribución proporcional por horas
# programadas — daba montos incorrectos (ej. $75 en vez de $120).
SQL_DETALLE = """
SELECT
    nq.docente_id,
    ndp.programa_id,
    p.nombre        AS programa_nombre,
    p.razon_social::text  AS razon_social,
    p.nivel,
    ndp.horas_presenciales,
    ndp.horas_virtuales,
    ndp.horas_suplencia,
    ndp.costo_hora,
    ndp.honorarios
FROM nomina_detalle_programa ndp
JOIN nomina_quincena nq ON nq.id = ndp.nomina_id
JOIN programas p         ON p.id = ndp.programa_id
WHERE nq.quincena_id = %s
  AND nq.docente_id  = ANY(%s)
ORDER BY nq.docente_id, p.nombre
"""


# ── Decimal helpers ────────────────────────────────────────────────────────────

def _d(v) -> Decimal:
    return Decimal("0") if v is None else Decimal(str(v))

def _r(v: Decimal) -> Decimal:
    return v.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


# ── Desglose por programa (datos reales) ─────────────────────────────────────

def _distribuir(nomina: dict, detalles: List[dict]) -> List[dict]:
    """
    Genera desglose por programa usando datos REALES de nomina_detalle_programa.
    Los descuentos y ajustes se distribuyen proporcional al peso de honorarios
    de cada programa (mucho más preciso que la distribución por horas programadas).
    """
    if not detalles:
        # Docente sin detalle — fallback (no debería pasar con cálculo correcto)
        return [{
            "programa_nombre":    "SIN ASIGNACIÓN",
            "razon_social":       str(nomina.get("adscripcion") or ""),
            "monto_presencial":   _d(nomina["honorarios"]),
            "monto_virtual":      Decimal("0"),
            "desc_presencial":    _d(nomina["descuentos_total"]),
            "desc_virtual":       Decimal("0"),
            "monto_otros":        _d(nomina["ajustes"]),
        }]

    dsc = _d(nomina["descuentos_total"])
    adj = _d(nomina["ajustes"])

    # Agrupar detalles por programa (pueden haber múltiples filas por programa
    # si el docente tiene varias asignaciones en el mismo programa)
    prog_map: Dict[str, dict] = {}
    for det in detalles:
        pnom = det["programa_nombre"]
        if pnom not in prog_map:
            prog_map[pnom] = {
                "programa_nombre": pnom,
                "razon_social":    str(det["razon_social"]),
                "hp": Decimal("0"), "hv": Decimal("0"),
                "costo_hora": _d(det["costo_hora"]),
                "hon": Decimal("0"),
            }
        prog_map[pnom]["hp"]  += _d(det["horas_presenciales"])
        prog_map[pnom]["hv"]  += _d(det["horas_virtuales"])
        prog_map[pnom]["hon"] += _d(det["honorarios"])

    hon_total = sum(p["hon"] for p in prog_map.values()) or Decimal("1")

    resultado = []
    es_clinico = lambda n: any(k in n.upper() for k in ("CAMPO", "CLINICO", "CLÍNICO"))

    for pnom, pd_ in prog_map.items():
        if es_clinico(pnom):
            resultado.append({
                "programa_nombre": pnom,
                "razon_social":    pd_["razon_social"],
                "monto_presencial": Decimal("0"),
                "monto_virtual":    Decimal("0"),
                "desc_presencial":  Decimal("0"),
                "desc_virtual":     Decimal("0"),
                "monto_otros":      Decimal("2500.00"),
            })
            continue

        prop = pd_["hon"] / hon_total
        mp   = _r(pd_["hp"] * pd_["costo_hora"])
        mv   = _r(pd_["hv"] * pd_["costo_hora"])
        resultado.append({
            "programa_nombre": pnom,
            "razon_social":    pd_["razon_social"],
            "monto_presencial": mp,
            "monto_virtual":    mv,
            "desc_presencial":  _r(dsc * prop),
            "desc_virtual":     Decimal("0"),
            "monto_otros":      _r(adj * prop),
        })

    return resultado


def _clasificar(nominas, detalle_por_docente):
    """Clasifica filas en centro_rows e instituto_map usando razon_social del programa."""
    centro_rows   = []
    instituto_map = {}
    for nom in nominas:
        desglose = _distribuir(nom, detalle_por_docente.get(nom["docente_id"], []))
        for dp in desglose:
            rs = str(dp.get("razon_social") or "").lower()
            if rs == "centro":
                centro_rows.append((nom, dp))
            elif rs == "instituto":
                instituto_map.setdefault(dp["programa_nombre"], []).append((nom, dp))
    return centro_rows, instituto_map


# ── Construcción hojas ─────────────────────────────────────────────────────────

COLS_CENTRO = [
    "PROGRAMA", "DOCENTE", "NOI",
    "PRESENCIAL ($)", "DESCUENTO", "AJUSTES",
    "HONORARIOS", "IVA 16%", "RET. ISR 10%", "RET. IVA", "TOTAL A PAGAR",
]
COLS_INSTITUTO = [
    "PROGRAMA", "DOCENTE", "NOI",
    "PRESENCIAL ($)", "VIRTUAL ($)", "DESCUENTO", "AJUSTES",
    "HONORARIOS", "IVA 16%", "RET. ISR 10%", "RET. IVA", "TOTAL A PAGAR",
]

ANCHOS_CENTRO   = {1:22, 2:32, 3:8, 4:14, 5:13, 6:12,
                   7:14, 8:11, 9:12, 10:10, 11:14}
ANCHOS_INSTITUTO = {1:22, 2:32, 3:8, 4:14, 5:13, 6:13, 7:12,
                    8:14, 9:11, 10:12, 11:10, 12:14}


def _escribir_hoja(ws, cols: List[str],
                   prog_order: List[str],
                   prog_rows: Dict[str, list],
                   anchos: Dict[int, int],
                   es_instituto: bool) -> None:
    n_cols   = len(cols)
    # CENTRO cols monetarias 4-11 · INSTITUTO cols monetarias 4-12
    cols_num = list(range(4, 13 if es_instituto else 12))

    # ── Encabezado ──
    for c, texto in enumerate(cols, 1):
        _hdr_cell(ws.cell(1, c), texto)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    fila     = 2
    fila_ini = 2

    for prog_idx, pnom in enumerate(prog_order):
        # Color específico por nombre de programa
        bg_sep, fg_sep, bg_par, bg_impar = _color_programa(pnom, prog_idx)

        filas = prog_rows[pnom]
        filas.sort(key=lambda x: x[0]["nombre_completo"])

        # ── Separador de programa ──
        _sep_cell(ws, fila, n_cols, pnom, bg_sep, fg_sep)
        ws.row_dimensions[fila].height = 18
        fila += 1

        # ── Filas del programa ──
        for row_idx, (nom, dp) in enumerate(filas):
            # Alternar entre bg_par y bg_impar dentro del mismo programa
            bg   = bg_par if row_idx % 2 == 0 else bg_impar
            mp   = dp["monto_presencial"]
            mv   = dp["monto_virtual"]
            dp_p = dp["desc_presencial"]
            dp_v = dp["desc_virtual"]
            otros= dp["monto_otros"]
            desc = dp_p + dp_v
            tot  = mp + mv + otros - desc          # honorarios netos (base fiscal)

            # ── Cálculo fiscal Art. 106 LISR ──────────────────────────────────
            iva_row     = _r(tot * Decimal("0.16"))
            ret_isr_row = _r(tot * Decimal("0.10"))
            ret_iva_row = _r(iva_row * Decimal("2") / Decimal("3"))
            total_pagar = _r(tot + iva_row - ret_isr_row - ret_iva_row)

            _data_cell(ws, fila, 1, pnom,                   bg=bg)
            _data_cell(ws, fila, 2, nom["nombre_completo"], bg=bg)
            _data_cell(ws, fila, 3, nom["noi"] or "",       bg=bg)

            if es_instituto:
                _data_cell(ws, fila, 4, float(mp),   bg=bg, money=True)
                _data_cell(ws, fila, 5, float(mv),   bg=bg, money=True)
                _data_cell(ws, fila, 6, float(desc), bg=bg, money=True, rojo=bool(desc))
                _data_cell(ws, fila, 7, float(otros),bg=bg, money=True)
                h, iv, r1, r2, tp = 8, 9, 10, 11, 12
            else:
                _data_cell(ws, fila, 4, float(mp),   bg=bg, money=True)
                _data_cell(ws, fila, 5, float(desc), bg=bg, money=True, rojo=bool(desc))
                _data_cell(ws, fila, 6, float(otros),bg=bg, money=True)
                h, iv, r1, r2, tp = 7, 8, 9, 10, 11

            _data_cell(ws, fila, h,  float(tot),         bg=bg, money=True)
            _data_cell(ws, fila, iv, float(iva_row),     bg=bg, money=True)
            _data_cell(ws, fila, r1, float(ret_isr_row), bg=bg, money=True,
                       rojo=bool(ret_isr_row))
            _data_cell(ws, fila, r2, float(ret_iva_row), bg=bg, money=True,
                       rojo=bool(ret_iva_row))
            _data_cell(ws, fila, tp, float(total_pagar), bg=bg, money=True)
            ws.row_dimensions[fila].height = 16
            fila += 1

    # ── Totales ──
    if fila > fila_ini:
        fin = fila - 1
        _tot_cell(ws, fila, 2, "TOTALES")
        for ci in cols_num:
            letra = get_column_letter(ci)
            _tot_cell(ws, fila, ci,
                      f"=SUM({letra}{fila_ini}:{letra}{fin})", money=True)
        ws.row_dimensions[fila].height = 18

    _auto_ancho(ws, anchos)


# ── Función pública ────────────────────────────────────────────────────────────

def generar_nomina_resumen_excel(conn, quincena_id: int) -> bytes:
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl no está instalado")

    cur = conn.cursor(cursor_factory=RealDictCursor)

    cur.execute(SQL_QUINCENA, (quincena_id,))
    quincena = cur.fetchone()
    if not quincena:
        raise ValueError(f"Quincena {quincena_id} no encontrada")
    ciclo = quincena["ciclo"] or ""

    rs = quincena["razon_social"]  # 'centro', 'instituto', 'ambas'

    cur.execute(SQL_NOMINA, (quincena_id,))
    nominas = [dict(r) for r in cur.fetchall()]
    if not nominas:
        raise ValueError(
            f"No hay nómina calculada para la quincena {quincena_id}. "
            "Ejecute el cálculo antes de exportar."
        )

    ids = list({n["docente_id"] for n in nominas})
    if ids:
        # Datos REALES per-programa desde nomina_detalle_programa
        cur.execute(SQL_DETALLE, (quincena_id, ids))
        detalle_por_doc = {}
        for row in cur.fetchall():
            detalle_por_doc.setdefault(row["docente_id"], []).append(dict(row))
    else:
        detalle_por_doc = {}
    cur.close()

    centro_rows, instituto_map = _clasificar(nominas, detalle_por_doc)

    # Ordenar programas CENTRO (por orden de aparición / alfabético)
    centro_prog_order = []
    centro_prog_rows  = {}
    for nom, dp in centro_rows:
        p = dp["programa_nombre"]
        if p not in centro_prog_rows:
            centro_prog_rows[p] = []
            centro_prog_order.append(p)
        centro_prog_rows[p].append((nom, dp))
    centro_prog_order.sort()

    # Ordenar programas INSTITUTO (orden canónico)
    def _peso(n):
        nu = n.upper()
        for i, pat in enumerate(ORDEN_INSTITUTO):
            if pat in nu or nu in pat:
                return i
        return 999
    instituto_prog_order = sorted(instituto_map.keys(), key=_peso)

    wb = openpyxl.Workbook()
    # wb siempre crea una hoja activa por defecto; la usamos para la primera
    # hoja necesaria y creamos la segunda solo si aplica.

    incluir_centro   = rs in ("centro", "ambas")
    incluir_instituto = rs in ("instituto", "ambas")

    primera_hoja = True

    if incluir_centro:
        ws_c = wb.active
        ws_c.title = "CENTRO"
        primera_hoja = False
        if centro_rows:
            _escribir_hoja(ws_c, COLS_CENTRO,
                           centro_prog_order, centro_prog_rows,
                           ANCHOS_CENTRO, es_instituto=False)
        else:
            ws_c.cell(1, 1, "Sin docentes en CENTRO para esta quincena")

    if incluir_instituto:
        if primera_hoja:
            ws_i = wb.active
            ws_i.title = "INSTITUTO"
            primera_hoja = False
        else:
            ws_i = wb.create_sheet("INSTITUTO")
        if instituto_map:
            instituto_prog_rows = {p: instituto_map[p] for p in instituto_prog_order}
            _escribir_hoja(ws_i, COLS_INSTITUTO,
                           instituto_prog_order, instituto_prog_rows,
                           ANCHOS_INSTITUTO, es_instituto=True)
        else:
            ws_i.cell(1, 1, "Sin docentes en INSTITUTO para esta quincena")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
