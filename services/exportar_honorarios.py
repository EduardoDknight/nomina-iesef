"""
Generación del Excel HONORARIOS CENTRO y HONORARIOS INSTITUTO.
Formato exacto del archivo actual:

Encabezado:
  A1: PROGRAMA ACADEMICO   I1: QUINCENA   N1: FECHA DE PAGO
  A2: [nombre programa]    I2: [período]  N2: [fecha] DE [mes] [año]
  N3: FORMA DE PAGO
  N4: HONORARIOS (o ASIMILADOS A SALARIOS)

Fila 7 (cabecera de tabla):
  PROGRAMA EDUCATIVO | NOMBRE | HORAS PROGRAMADAS | HORAS PRESENCIALES |
  HORAS VIRTUALES | DESCUENTOS | COSTO POR HORA | HONORARIOS | IVA 16% |
  SUB-TOTAL | RETENCION ISR | RETENCION IVA | TOTAL A PAGAR | FIRMA

Filas de datos: una por docente (o por programa si da en varios)
"""
import io
from decimal import Decimal
from datetime import date
from typing import List
import psycopg2
from psycopg2.extras import RealDictCursor

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Meses en español
MESES = {
    1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
    5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
    9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
}

def _estilo_encabezado():
    return Font(bold=True, size=11)

def _estilo_cabecera_tabla():
    return Font(bold=True, size=10, color="FFFFFF")

def _fill_cabecera():
    return PatternFill("solid", fgColor="2F4F8F")

def _border_thin():
    lado = Side(style="thin")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def _formato_quincena(fecha_inicio: date, fecha_fin: date) -> str:
    """Ej: '11-25 DE FEBRERO\n2ª QUINCENA'"""
    mes = MESES[fecha_inicio.month]
    num = "1ª" if fecha_inicio.day <= 15 else "2ª"
    return f"{fecha_inicio.day}-{fecha_fin.day} DE {mes}\n{num} QUINCENA"


def generar_honorarios_excel(
    conn,
    quincena_id: int,
    razon_social: str,   # 'centro' o 'instituto'
    fecha_pago: date
) -> bytes:
    """
    Genera el Excel HONORARIOS para la razón social indicada.
    Retorna bytes del archivo .xlsx listo para descarga.
    """
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl no está instalado")

    cur = conn.cursor(cursor_factory=RealDictCursor)

    # ── Obtener quincena ───────────────────────────────────────────────────────
    cur.execute("SELECT * FROM quincenas WHERE id = %s", (quincena_id,))
    quincena = cur.fetchone()
    if not quincena:
        raise ValueError(f"Quincena {quincena_id} no encontrada")

    # ── Obtener nóminas de docentes de esta razón social ─────────────────────
    if razon_social == "centro":
        adscripcion_filter = ("centro", "ambos")
    else:
        adscripcion_filter = ("instituto", "ambos")

    cur.execute("""
        SELECT
            nq.*,
            d.nombre_completo,
            d.adscripcion,
            d.regimen_fiscal,
            d.costo_hora_centro,
            d.costo_hora_instituto
        FROM nomina_quincena nq
        JOIN docentes d ON nq.docente_id = d.id
        WHERE nq.quincena_id = %s
          AND d.adscripcion = ANY(%s)
          AND nq.total_final > 0
        ORDER BY d.nombre_completo
    """, (quincena_id, list(adscripcion_filter)))
    nominas = cur.fetchall()

    # ── Obtener detalle por programa para cada docente ────────────────────────
    detalles = {}
    if nominas:
        nomina_ids = [n["id"] for n in nominas]
        cur.execute("""
            SELECT ndp.*, p.nombre AS programa_nombre, p.razon_social AS prog_razon
            FROM nomina_detalle_programa ndp
            JOIN programas p ON ndp.programa_id = p.id
            WHERE ndp.nomina_id = ANY(%s)
            ORDER BY ndp.nomina_id, p.nombre
        """, (nomina_ids,))
        for row in cur.fetchall():
            detalles.setdefault(row["nomina_id"], []).append(row)

    cur.close()

    # ── Crear workbook ─────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"HONORARIOS {razon_social.upper()}"

    # Configurar anchos de columna (A-N)
    anchos = [22, 28, 8, 8, 8, 6, 7, 10, 9, 10, 9, 9, 10, 12]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    # ── ENCABEZADO ────────────────────────────────────────────────────────────
    if razon_social == "centro":
        programas_titulo = ["PREPARATORIA"]
    else:
        programas_titulo = [
            "LICENCIATURA EN NUTRICIÓN",
            "LICENCIATURA EN ENFERMERÍA",
            "LIC. ENFERMERÍA NIVELACIÓN ACADÉMICA",
            "ESPECIALIDADES",
            "MAESTRÍAS",
        ]

    fila_prog = 1
    for prog in programas_titulo:
        ws.cell(fila_prog, 1, prog).font = _estilo_encabezado()
        fila_prog += 1

    # Quincena
    ws.cell(1, 9, "QUINCENA").font = _estilo_encabezado()
    ws.cell(2, 9, _formato_quincena(quincena["fecha_inicio"], quincena["fecha_fin"]))
    ws.cell(2, 9).alignment = Alignment(wrap_text=True)

    # Fecha de pago
    ws.cell(1, 14, "FECHA DE PAGO").font = _estilo_encabezado()
    ws.cell(2, 14, str(fecha_pago.day))
    ws.cell(2, 15, "DE")
    ws.cell(2, 16, MESES[fecha_pago.month])
    ws.cell(2, 17, str(fecha_pago.year))

    ws.cell(3, 14, "FORMA DE PAGO").font = _estilo_encabezado()
    ws.cell(4, 14, "HONORARIOS").font = _estilo_encabezado()

    # ── FILA DE CABECERA DE TABLA (fila 7) ────────────────────────────────────
    CABECERAS = [
        "PROGRAMA EDUCATIVO", "NOMBRE",
        "HORAS PROGRAMADAS", "HORAS PRESENCIALES", "HORAS VIRTUALES",
        "DESCUENTOS", "COSTO POR HORA",
        "HONORARIOS", "IVA 16%", "SUB-TOTAL",
        "RETENCION ISR", "RETENCION IVA", "TOTAL A PAGAR", "FIRMA"
    ]
    FILA_CAB = 7
    for col, cab in enumerate(CABECERAS, 1):
        cell = ws.cell(FILA_CAB, col, cab)
        cell.font      = _estilo_cabecera_tabla()
        cell.fill      = _fill_cabecera()
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border    = _border_thin()
    ws.row_dimensions[FILA_CAB].height = 30

    # ── FILAS DE DATOS ────────────────────────────────────────────────────────
    fila = FILA_CAB + 1
    primer_docente_fiscal = True  # el primer docente lleva el cálculo fiscal

    for nomina in nominas:
        det = detalles.get(nomina["id"], [])

        # Filtrar solo programas de esta razón social
        if razon_social == "centro":
            det_rs = [d for d in det if d["prog_razon"] in ("centro", "ambos")]
        else:
            det_rs = [d for d in det if d["prog_razon"] in ("instituto", "ambos")]

        # Si no hay detalle, usar la nómina completa en una sola fila
        if not det_rs:
            det_rs = [None]

        for i, dp in enumerate(det_rs):
            es_primera = (i == 0)
            prog_nombre = dp["programa_nombre"] if dp else ""
            horas_pres  = float(dp["horas_presenciales"] or 0) if dp else float(nomina["horas_presenciales"] or 0)
            horas_virt  = float(dp["horas_virtuales"]    or 0) if dp else float(nomina["horas_virtuales"]    or 0)
            descuentos  = float(nomina["horas_descuento"] or 0) if es_primera else 0
            costo_hora  = float(dp["costo_hora"]         or 0) if dp else 0

            ws.cell(fila, 1, prog_nombre)
            ws.cell(fila, 2, nomina["nombre_completo"] if es_primera else "")
            ws.cell(fila, 3, float(nomina["horas_programadas"] or 0) if es_primera else "")
            ws.cell(fila, 4, horas_pres)
            ws.cell(fila, 5, horas_virt)
            ws.cell(fila, 6, descuentos)
            ws.cell(fila, 7, costo_hora)

            # El cálculo fiscal solo va en la PRIMERA fila del docente
            if es_primera:
                ws.cell(fila, 8,  float(nomina["honorarios"]    or 0))
                ws.cell(fila, 9,  float(nomina["iva"]           or 0))
                ws.cell(fila, 10, float(nomina["sub_total"]     or 0))
                ws.cell(fila, 11, float(nomina["retencion_isr"] or 0))
                ws.cell(fila, 12, float(nomina["retencion_iva"] or 0))
                ws.cell(fila, 13, float(nomina["total_final"]   or 0))

            # Formato monetario
            for col in (8, 9, 10, 11, 12, 13):
                c = ws.cell(fila, col)
                c.number_format = '#,##0.00'

            # Bordes
            for col in range(1, 15):
                ws.cell(fila, col).border = _border_thin()

            # Resaltar fila del docente (primera)
            if es_primera:
                for col in range(1, 15):
                    ws.cell(fila, col).fill = PatternFill("solid", fgColor="EEF2FF")

            fila += 1

    # ── FILA DE TOTALES ───────────────────────────────────────────────────────
    fila_inicio_datos = FILA_CAB + 1
    fila_fin_datos    = fila - 1
    if fila_fin_datos >= fila_inicio_datos:
        ws.cell(fila, 2, "TOTALES").font = Font(bold=True)
        for col, letra_col in [
            (8, get_column_letter(8)), (9, get_column_letter(9)),
            (10, get_column_letter(10)), (11, get_column_letter(11)),
            (12, get_column_letter(12)), (13, get_column_letter(13))
        ]:
            c = ws.cell(fila, col,
                f"=SUM({letra_col}{fila_inicio_datos}:{letra_col}{fila_fin_datos})")
            c.number_format = '#,##0.00'
            c.font = Font(bold=True)
            c.border = _border_thin()

    # ── Exportar a bytes ───────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
