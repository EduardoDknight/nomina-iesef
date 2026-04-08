from fastapi import APIRouter, HTTPException, Depends, Request
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime
import psycopg2
from psycopg2.extras import RealDictCursor
import bcrypt
import logging

from config import settings
from routers.auth import get_usuario_actual, UsuarioActual
from services.auditoria import registrar, ip_from_request

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/usuarios", tags=["usuarios"])

ROLES_VALIDOS = [
    'director_cap_humano', 'cap_humano', 'finanzas',
    'coord_docente', 'servicios_escolares',
    'coord_academica', 'educacion_virtual',
    'docente', 'reportes',
]

def get_conn():
    return psycopg2.connect(settings.database_url_nomina, cursor_factory=RealDictCursor)

# ── Modelos ────────────────────────────────────────────────────────────────────

class UsuarioOut(BaseModel):
    id: int
    nombre: str
    email: str
    rol: str
    programa_id: Optional[int] = None
    activo: bool
    ultimo_acceso: Optional[datetime] = None
    creado_en: datetime

class UsuarioCreate(BaseModel):
    nombre: str
    email: str
    rol: str
    password: str = "IESEF2026"
    programa_id: Optional[int] = None

class UsuarioUpdate(BaseModel):
    nombre:      Optional[str]  = None
    email:       Optional[str]  = None
    rol:         Optional[str]  = None
    programa_id: Optional[int]  = None
    activo:      Optional[bool] = None
    password:    Optional[str]  = None  # si se envía, se restablece

# ── Endpoints ──────────────────────────────────────────────────────────────────

@router.get("", response_model=List[UsuarioOut])
async def listar_usuarios(usuario: UsuarioActual = Depends(get_usuario_actual)):
    if usuario.rol not in ("director_cap_humano", "cap_humano"):
        raise HTTPException(status_code=403, detail="Sin permiso")
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, nombre, email, rol, programa_id, activo, ultimo_acceso, creado_en
        FROM usuarios
        ORDER BY
            CASE rol
                WHEN 'director_cap_humano' THEN 1
                WHEN 'cap_humano'          THEN 2
                WHEN 'finanzas'            THEN 3
                WHEN 'coord_docente'       THEN 4
                ELSE 5
            END, nombre
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return [UsuarioOut(**r) for r in rows]

@router.post("", response_model=UsuarioOut, status_code=201)
async def crear_usuario(
    request: Request,
    body: UsuarioCreate,
    usuario: UsuarioActual = Depends(get_usuario_actual)
):
    if usuario.rol != "director_cap_humano":
        raise HTTPException(status_code=403, detail="Solo el Director de Capital Humano puede crear usuarios")
    if body.rol not in ROLES_VALIDOS:
        raise HTTPException(status_code=400, detail=f"Rol inválido")
    pwd_hash = bcrypt.hashpw(body.password.encode(), bcrypt.gensalt(rounds=10)).decode()
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO usuarios (nombre, email, password_hash, rol, programa_id, activo)
            VALUES (%s, %s, %s, %s, %s, true)
            RETURNING id, nombre, email, rol, programa_id, activo, ultimo_acceso, creado_en
        """, (body.nombre, body.email, pwd_hash, body.rol, body.programa_id))
        row = cur.fetchone()
        registrar(conn, usuario, accion="crear_usuario", entidad="usuario",
                  entidad_id=str(row["id"]),
                  descripcion=f"Creó usuario {body.email} ({body.rol})",
                  detalle={"email": body.email, "rol": body.rol},
                  ip=ip_from_request(request))
        conn.commit()
        cur.close()
        conn.close()
        return UsuarioOut(**row)
    except psycopg2.errors.UniqueViolation:
        raise HTTPException(status_code=409, detail="El correo ya está registrado")
    except Exception as e:
        logger.error(f"Error crear_usuario: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.patch("/{usuario_id}", response_model=UsuarioOut)
async def actualizar_usuario(
    usuario_id: int,
    body: UsuarioUpdate,
    request: Request,
    usuario: UsuarioActual = Depends(get_usuario_actual)
):
    if usuario.rol != "director_cap_humano":
        raise HTTPException(status_code=403, detail="Solo el Director de Capital Humano puede modificar usuarios")
    if body.rol and body.rol not in ROLES_VALIDOS:
        raise HTTPException(status_code=400, detail="Rol inválido")

    updates = {k: v for k, v in body.model_dump(exclude_none=True).items() if k != "password"}
    if body.password:
        updates["password_hash"] = bcrypt.hashpw(body.password.encode(), bcrypt.gensalt(rounds=10)).decode()
    if not updates:
        raise HTTPException(status_code=400, detail="Sin campos para actualizar")

    set_clause = ", ".join(f"{k} = %s" for k in updates)
    values = list(updates.values()) + [usuario_id]
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            f"UPDATE usuarios SET {set_clause} WHERE id = %s "
            "RETURNING id, nombre, email, rol, programa_id, activo, ultimo_acceso, creado_en",
            values
        )
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")
        registrar(conn, usuario, accion="actualizar_usuario", entidad="usuario",
                  entidad_id=str(usuario_id),
                  descripcion=f"Actualizó usuario {row['email']}",
                  detalle={"campos": list(updates.keys())},
                  ip=ip_from_request(request))
        conn.commit()
        cur.close()
        conn.close()
        return UsuarioOut(**row)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error actualizar_usuario: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ── Credenciales de acceso ─────────────────────────────────────────────────────

PUEDE_VER_CREDS_DOCENTES     = ("superadmin", "director_cap_humano", "cap_humano", "coord_docente")
PUEDE_VER_CREDS_TRABAJADORES = ("superadmin", "director_cap_humano", "cap_humano")
PUEDE_RESETEAR               = ("superadmin", "director_cap_humano")

import io
from datetime import datetime as _dt


@router.get("/credenciales-docentes")
async def credenciales_docentes(usuario: UsuarioActual = Depends(get_usuario_actual)):
    if usuario.rol not in PUEDE_VER_CREDS_DOCENTES:
        raise HTTPException(status_code=403, detail="Sin permiso")
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT u.id, u.nombre, u.email AS username, u.debe_cambiar_password,
                   u.ultimo_acceso, u.activo,
                   d.numero_docente
            FROM usuarios u
            JOIN docentes d ON d.id = u.docente_id
            WHERE u.rol = 'docente' AND u.docente_id IS NOT NULL
            ORDER BY u.nombre
        """)
        rows = cur.fetchall()
        return [
            {
                "id":                    r["id"],
                "nombre":                r["nombre"],
                "username":              r["username"],
                "numero_docente":        r["numero_docente"],
                "debe_cambiar_password": r["debe_cambiar_password"],
                "password_visible":      f"IESEF{_dt.now().year}" if r["debe_cambiar_password"] else "••••••",
                "ultimo_acceso":         r["ultimo_acceso"],
                "activo":                r["activo"],
            }
            for r in rows
        ]
    finally:
        cur.close()
        conn.close()


@router.get("/credenciales-trabajadores")
async def credenciales_trabajadores(usuario: UsuarioActual = Depends(get_usuario_actual)):
    if usuario.rol not in PUEDE_VER_CREDS_TRABAJADORES:
        raise HTTPException(status_code=403, detail="Sin permiso")
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT u.id, u.nombre, u.email AS username, u.debe_cambiar_password,
                   u.ultimo_acceso, u.activo,
                   t.chec_id, t.cargo
            FROM usuarios u
            JOIN trabajadores t ON t.id = u.trabajador_id
            WHERE u.rol = 'trabajador' AND u.trabajador_id IS NOT NULL
            ORDER BY u.nombre
        """)
        rows = cur.fetchall()
        return [
            {
                "id":                    r["id"],
                "nombre":                r["nombre"],
                "username":              r["username"],
                "chec_id":               r["chec_id"],
                "cargo":                 r["cargo"],
                "debe_cambiar_password": r["debe_cambiar_password"],
                "password_visible":      f"IESEF{_dt.now().year}" if r["debe_cambiar_password"] else "••••••",
                "ultimo_acceso":         r["ultimo_acceso"],
                "activo":                r["activo"],
            }
            for r in rows
        ]
    finally:
        cur.close()
        conn.close()


@router.post("/{usuario_id}/reset-password", status_code=200)
async def reset_password(
    usuario_id: int,
    usuario: UsuarioActual = Depends(get_usuario_actual)
):
    if usuario.rol not in PUEDE_RESETEAR:
        raise HTTPException(status_code=403, detail="Solo superadmin o Director de Capital Humano")
    password_default = f"IESEF{_dt.now().year}"
    nuevo_hash = bcrypt.hashpw(password_default.encode(), bcrypt.gensalt(rounds=10)).decode()
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            UPDATE usuarios
            SET password_hash = %s, debe_cambiar_password = true
            WHERE id = %s AND rol IN ('docente', 'trabajador')
            RETURNING id, nombre, email
        """, (nuevo_hash, usuario_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Usuario no encontrado o no es docente/trabajador")
        conn.commit()
        return {"ok": True, "nombre": row["nombre"], "password_reset_a": password_default}
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"Error reset_password: {e}")
        raise HTTPException(status_code=500, detail="Error interno")
    finally:
        cur.close()
        conn.close()


@router.get("/credenciales-docentes/export")
async def export_credenciales_docentes(usuario: UsuarioActual = Depends(get_usuario_actual)):
    """Exporta Excel con usuarios docentes."""
    if usuario.rol not in PUEDE_VER_CREDS_DOCENTES:
        raise HTTPException(status_code=403, detail="Sin permiso")
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT u.nombre, u.email AS username, u.debe_cambiar_password,
                   u.ultimo_acceso, d.numero_docente
            FROM usuarios u
            JOIN docentes d ON d.id = u.docente_id
            WHERE u.rol = 'docente' AND u.docente_id IS NOT NULL
            ORDER BY u.nombre
        """)
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Credenciales Docentes"
    header_fill = PatternFill("solid", fgColor="1E40AF")
    header_font = Font(color="FFFFFF", bold=True)
    headers = ["No. Docente", "Nombre", "Usuario (correo)", "Contraseña", "Estado", "Último acceso"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_num, r in enumerate(rows, 2):
        ws.cell(row=row_num, column=1, value=r["numero_docente"])
        ws.cell(row=row_num, column=2, value=r["nombre"])
        ws.cell(row=row_num, column=3, value=r["username"])
        pwd = f"IESEF{_dt.now().year}" if r["debe_cambiar_password"] else "••••••"
        ws.cell(row=row_num, column=4, value=pwd)
        ws.cell(row=row_num, column=5, value="Pendiente cambio" if r["debe_cambiar_password"] else "Personalizada")
        ws.cell(row=row_num, column=6, value=str(r["ultimo_acceso"]) if r["ultimo_acceso"] else "Nunca")

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 24

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"credenciales_docentes_{_dt.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/credenciales-trabajadores/export")
async def export_credenciales_trabajadores(usuario: UsuarioActual = Depends(get_usuario_actual)):
    """Exporta Excel con usuarios trabajadores."""
    if usuario.rol not in PUEDE_VER_CREDS_TRABAJADORES:
        raise HTTPException(status_code=403, detail="Sin permiso")
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT u.nombre, u.email AS username, t.chec_id, t.cargo,
                   u.debe_cambiar_password, u.ultimo_acceso
            FROM usuarios u
            JOIN trabajadores t ON t.id = u.trabajador_id
            WHERE u.rol = 'trabajador' AND u.trabajador_id IS NOT NULL
            ORDER BY u.nombre
        """)
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Credenciales Trabajadores"
    header_fill = PatternFill("solid", fgColor="065F46")
    header_font = Font(color="FFFFFF", bold=True)
    headers = ["Chec ID", "Nombre", "Usuario", "Cargo", "Contraseña", "Estado", "Último acceso"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_num, r in enumerate(rows, 2):
        ws.cell(row=row_num, column=1, value=r["chec_id"])
        ws.cell(row=row_num, column=2, value=r["nombre"])
        ws.cell(row=row_num, column=3, value=r["username"])
        ws.cell(row=row_num, column=4, value=r["cargo"] or "")
        pwd = f"IESEF{_dt.now().year}" if r["debe_cambiar_password"] else "••••••"
        ws.cell(row=row_num, column=5, value=pwd)
        ws.cell(row=row_num, column=6, value="Pendiente cambio" if r["debe_cambiar_password"] else "Personalizada")
        ws.cell(row=row_num, column=7, value=str(r["ultimo_acceso"]) if r["ultimo_acceso"] else "Nunca")

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"credenciales_trabajadores_{_dt.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
