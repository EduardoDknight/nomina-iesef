"""
import_incidencias_excel.py
==========================
Importa el archivo INCIDENCIAS DOCENTE CICLO ESCOLAR FEBRERO-JULIO2026.xlsx
a la tabla incidencias de iesef_nomina.

Mapeo de hojas → programas:
  LICENCIATURA EN ENFERMERÍA  → programa_id = 2 (o busca por nombre)
  LICENCIATURA EN NUTRICIÓN   → programa_id = 3
  ESPECIALIDADES              → programa_id = 5
  LENA                        → programa_id = 4
  IDIOMAS                     → (si existe)
  PREPARATORIA                → programa_id = 1

Columnas del Excel (fila 3 = headers, fila 4+ = datos):
  A=Fecha, B=Grupo, C=Horario, D=Materia, E=Docente titular,
  F=Docente sustituto, G=Observaciones

USO:
  python scripts/import_incidencias_excel.py [--dry-run]
"""
import sys, re
from datetime import date, datetime
import psycopg2
from psycopg2.extras import RealDictCursor
import openpyxl
from difflib import SequenceMatcher

sys.stdout.reconfigure(encoding='utf-8')

DRY_RUN = '--dry-run' in sys.argv

ARCHIVO = r'C:\Users\Admin\Downloads\INCIDENCIAS DOCENTE CICLO ESCOLAR FEBRERO-JULIO2026.xlsx'
CONN_STR = 'postgresql://nomina_user:IESEFnomina%402026%24@localhost:5432/iesef_nomina'
QUINCENA_ID = 3   # quincena Mar 11-25 — ajustar si se importa otra quincena

# Hojas de programa a procesar (la última fecha de fecha_inicio de la quincena decide qué incluir)
HOJAS_PROGRAMA = [
    'LICENCIATURA EN ENFERMERÍA',
    'LICENCIATURA EN NUTRICIÓN',
    'ESPECIALIDADES',
    'LENA',
    'IDIOMAS',
    'PREPARATORIA',
]

# ── helpers ────────────────────────────────────────────────────────────────────

def normalizar(nombre: str) -> str:
    """Normaliza nombre para comparación: minúsculas, sin tildes, sin grado académico."""
    if not nombre:
        return ''
    nombre = nombre.strip()
    # Quitar títulos académicos al inicio (tokens que terminan en "." y son abreviaturas):
    # maneja: "Dr.", "Dra.", "Mtra.", "Mtro.", "Lic.", "Ing.", "Ing. Biom.", "L.N.", "L.E.E.",
    # "M.C.B.S.", "M.N.H.", "M.D.E.M.S.", "Med. Cir.", etc.
    # Estrategia: eliminar en loop cualquier token inicial que parezca abreviatura (no apellido)
    for _ in range(6):  # máx 6 prefijos
        m = re.match(r'^([A-ZÁÉÍÓÚÜ][a-záéíóúüA-ZÁÉÍÓÚÜ]*\.(?:\s*[A-ZÁÉÍÓÚÜ][a-záéíóúüA-ZÁÉÍÓÚÜ]*\.)*)\s+', nombre)
        if m:
            nombre = nombre[m.end():]
        else:
            break
    # Normalizar tildes y pasar a minúsculas
    nombre = nombre.lower()
    repl = {'á':'a','é':'e','í':'i','ó':'o','ú':'u','ü':'u','ñ':'n'}
    for k,v in repl.items():
        nombre = nombre.replace(k,v)
    # Ordenar palabras alfabéticamente para que el orden nombre/apellido no importe
    # "Luis Alberto Lazcano Aguilar" == "Lazcano Aguilar Luis Alberto" → mismo resultado
    words = sorted(nombre.strip().split())
    return ' '.join(words)

def similitud(a: str, b: str) -> float:
    return SequenceMatcher(None, normalizar(a), normalizar(b)).ratio()

def buscar_docente(nombre_excel: str, docentes: list) -> dict | None:
    """Busca el docente más similar con umbral 0.65."""
    if not nombre_excel or not nombre_excel.strip():
        return None
    mejor = None
    mejor_score = 0.0
    n = normalizar(nombre_excel)
    for d in docentes:
        s = similitud(n, d['nombre_completo'])
        if s > mejor_score:
            mejor_score = s
            mejor = d
    if mejor_score >= 0.65:
        return mejor
    return None

def buscar_asignacion(docente_id: int, grupo: str, materia_nombre: str,
                      asignaciones: list) -> dict | None:
    """Busca la asignación del docente que mejor coincide con grupo y materia."""
    candidatos = [a for a in asignaciones if a['docente_id'] == docente_id]
    if not candidatos:
        return None
    if grupo:
        # Intentar match por grupo exacto primero
        por_grupo = [a for a in candidatos if
                     a.get('grupo','').strip().lower() == grupo.strip().lower()]
        if por_grupo:
            candidatos = por_grupo
    if materia_nombre:
        mejor = max(candidatos,
                    key=lambda a: similitud(a.get('materia_nombre',''), materia_nombre))
        if similitud(mejor.get('materia_nombre',''), materia_nombre) >= 0.55:
            return mejor
    # Si no matchea materia, tomar primera por grupo
    return candidatos[0] if candidatos else None

def parse_fecha(valor) -> date | None:
    if isinstance(valor, (date, datetime)):
        d = valor.date() if isinstance(valor, datetime) else valor
        # Corregir años erróneos (2008, 2025) → asumir 2026
        if d.year not in (2025, 2026):
            d = d.replace(year=2026)
        return d
    if isinstance(valor, str) and re.search(r'\d', valor):
        # intentar parsear texto
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
            try:
                return datetime.strptime(valor.strip(), fmt).date()
            except ValueError:
                pass
    return None

# ── main ──────────────────────────────────────────────────────────────────────

conn = psycopg2.connect(CONN_STR, cursor_factory=RealDictCursor)
cur = conn.cursor()

# Cargar catálogos
cur.execute("SELECT id, nombre_completo FROM docentes WHERE activo = true")
docentes_db = list(cur.fetchall())

cur.execute("""
    SELECT a.id, a.docente_id, a.grupo, a.ciclo,
           mat.nombre AS materia_nombre, p.nombre AS programa_nombre
    FROM asignaciones a
    JOIN materias mat ON a.materia_id = mat.id
    JOIN programas p  ON mat.programa_id = p.id
    WHERE a.activa = true
""")
asignaciones_db = list(cur.fetchall())

# Verificar quincena
cur.execute("SELECT id, ciclo, fecha_inicio, fecha_fin FROM quincenas WHERE id = %s", (QUINCENA_ID,))
quincena = cur.fetchone()
if not quincena:
    print(f"ERROR: quincena_id={QUINCENA_ID} no existe")
    sys.exit(1)
print(f"Quincena: {quincena['fecha_inicio']} → {quincena['fecha_fin']} (ciclo {quincena['ciclo']})")

wb = openpyxl.load_workbook(ARCHIVO, data_only=True)
print(f"Hojas disponibles: {wb.sheetnames}\n")

total_ok = 0
total_skip = 0
total_warn = 0

for hoja_nombre in HOJAS_PROGRAMA:
    if hoja_nombre not in wb.sheetnames:
        print(f"[SKIP] Hoja '{hoja_nombre}' no encontrada en el Excel")
        continue

    ws = wb[hoja_nombre]
    print(f"\n{'='*60}")
    print(f"Hoja: {hoja_nombre}")
    print(f"{'='*60}")

    # Encontrar fila de headers (contiene 'Fecha' en columna A)
    header_row = None
    for r in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        if r[0] and str(r[0]).strip().lower() == 'fecha':
            header_row = r
            break

    insertados = 0
    skipped = 0
    warnings = []
    fecha_actual = None  # para filas sin fecha (heredan de la anterior)

    for row in ws.iter_rows(min_row=4, values_only=True):
        col_fecha, col_grupo, col_horario, col_materia, col_titular, col_suplente, col_obs = \
            (row[i] if i < len(row) else None for i in range(7))

        # Separadores de semana (texto en col A)
        if isinstance(col_fecha, str) and re.search(r'semana|SEMANA', col_fecha, re.I):
            continue
        if col_fecha is None and col_titular is None and col_materia is None:
            continue  # fila vacía

        # Fecha
        f = parse_fecha(col_fecha)
        if f:
            fecha_actual = f
        elif fecha_actual is None:
            continue  # no tenemos fecha aún
        fecha = fecha_actual

        # Filtrar por rango de quincena
        if not (quincena['fecha_inicio'] <= fecha <= quincena['fecha_fin']):
            skipped += 1
            continue

        if not col_titular:
            skipped += 1
            continue

        # Buscar docente titular
        doc_titular = buscar_docente(str(col_titular), docentes_db)
        if not doc_titular:
            warnings.append(f"  ⚠ Titular no encontrado: '{col_titular}'  fecha={fecha}")
            total_warn += 1
            skipped += 1
            continue

        # Buscar asignación
        asig = buscar_asignacion(doc_titular['id'], str(col_grupo or ''),
                                 str(col_materia or ''), asignaciones_db)
        if not asig:
            warnings.append(f"  ⚠ Sin asignación: {doc_titular['nombre_completo']} | "
                          f"grupo={col_grupo} | materia={col_materia}")
            total_warn += 1
            skipped += 1
            continue

        # Determinar tipo
        tipo = 'suplencia' if col_suplente and str(col_suplente).strip() else 'falta'

        # Buscar suplente
        doc_suplente = None
        horas_suplidas = None
        if tipo == 'suplencia' and col_suplente:
            suplente_texto = str(col_suplente)
            # Extraer horas mencionadas en el texto del suplente
            m = re.search(r'cubrió\s+(\d+\.?\d*)\s+hora', suplente_texto, re.I)
            if m:
                horas_suplidas = float(m.group(1))
            doc_suplente = buscar_docente(suplente_texto, docentes_db)

        # Horas afectadas: inferir del horario si es posible
        horas_afectadas = 0
        if col_horario:
            # parsear formatos: "15:00 a 18:00", "14 - 16", "09:00-11:00"
            nums = re.findall(r'\d+(?:\.\d+)?', str(col_horario))
            if len(nums) >= 2:
                try:
                    h1, h2 = float(nums[0]), float(nums[1])
                    # Si son horas (0-24), calcular diferencia
                    if 0 <= h1 <= 24 and 0 <= h2 <= 24 and h2 > h1:
                        horas_afectadas = h2 - h1
                    elif len(nums) >= 4:  # ej "15 00 18 00"
                        h1 = float(nums[0]) + float(nums[1])/60
                        h2 = float(nums[2]) + float(nums[3])/60
                        horas_afectadas = max(0, h2 - h1)
                except:
                    pass

        if horas_afectadas == 0:
            # Fallback: horas del bloque de la asignación
            cur.execute("""
                SELECT COALESCE(horas_bloque, (EXTRACT(EPOCH FROM (hora_fin - hora_inicio))/3600)::numeric)
                FROM horario_clases
                WHERE asignacion_id = %s
                LIMIT 1
            """, (asig['id'],))
            r2 = cur.fetchone()
            if r2:
                horas_afectadas = float(list(r2.values())[0] or 0)

        obs = str(col_obs) if col_obs else (str(col_suplente) if col_suplente else None)

        print(f"  {'[DRY]' if DRY_RUN else '[INS]'} "
              f"{fecha} | {tipo:9} | {doc_titular['nombre_completo'][:30]:30} "
              f"| {str(col_materia or '')[:25]:25} | {horas_afectadas}h"
              + (f" → suplente: {doc_suplente['nombre_completo'][:20]}" if doc_suplente else ""))

        if not DRY_RUN:
            try:
                cur.execute("""
                    INSERT INTO incidencias
                        (quincena_id, docente_titular_id, asignacion_id, tipo, fecha,
                         horas_afectadas, docente_suplente_id, horas_suplidas,
                         estado, notas)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'pendiente',%s)
                    ON CONFLICT DO NOTHING
                """, (
                    QUINCENA_ID,
                    doc_titular['id'],
                    asig['id'],
                    tipo,
                    fecha,
                    horas_afectadas,
                    doc_suplente['id'] if doc_suplente else None,
                    horas_suplidas,
                    obs,
                ))
            except Exception as e:
                warnings.append(f"  ERROR INSERT: {e}")
                total_warn += 1
                skipped += 1
                conn.rollback()
                continue

        insertados += 1
        total_ok += 1

    for w in warnings:
        print(w)
    print(f"  → {insertados} insertadas, {skipped} omitidas, {len(warnings)} advertencias")

# commit
if not DRY_RUN:
    conn.commit()
    print(f"\n✅ COMMIT — Total insertadas: {total_ok} | Omitidas: {total_skip} | Advertencias: {total_warn}")
else:
    print(f"\n[DRY RUN] Sin cambios. Total que se insertarían: {total_ok} | Advertencias: {total_warn}")

cur.close()
conn.close()
