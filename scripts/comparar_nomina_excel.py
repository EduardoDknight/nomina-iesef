"""
Compara la nómina calculada por el sistema nuevo vs el Excel del sistema v1.
Quincena 3 — Mar 11-25, 2026

Salida: tabla de diferencias por docente/programa.
"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import psycopg2
from psycopg2.extras import RealDictCursor
from decimal import Decimal
from difflib import SequenceMatcher
import openpyxl

# ── Config ─────────────────────────────────────────────────────────────────────
DB_URL    = 'postgresql://nomina_user:IESEFnomina%402026%24@localhost:5432/iesef_nomina'
EXCEL     = r'C:\Users\Admin\Downloads\NOMINA 2DA DE MARZO-nomina final que entrega el nuevo sistema v1 local ya con correcciones del personal académnico, virtual y docentes, docentes de practicas.xlsx'
QUINCENA_ID   = 3
FECHA_INICIO  = '2026-03-11'
FECHA_FIN     = '2026-03-25'

# Mapeo hoja Excel → programa_id en DB
HOJA_PROGRAMA = {
    'PREPA':          1,
    'ENFERMERIA':     2,
    'NUTRICION':      3,
    'LENA':           4,
    'ESPECIALIDADES': 5,
    'MAESTRIAS':      6,
    'CAMPO CLINICO':  7,
}

# ── Helpers ────────────────────────────────────────────────────────────────────

def sim(a, b):
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()

def normalizar(nombre):
    if not nombre:
        return ''
    prefijos = ['dra.','dr.','mtra.','mtro.','lic.','enf.','lic','dra','dr','mtra','mtro']
    palabras = nombre.strip().split()
    filtradas = [p for p in palabras if p.lower().rstrip('.') not in prefijos]
    return ' '.join(sorted(filtradas, key=str.lower)).lower()

def buscar_docente_db(nombre_excel, docentes_db):
    """Devuelve el docente_id de la mejor coincidencia."""
    ne = normalizar(nombre_excel)
    mejor_id, mejor_score = None, 0.0
    for doc in docentes_db:
        score = sim(ne, normalizar(doc['nombre_completo']))
        if score > mejor_score:
            mejor_score = score
            mejor_id = doc['id']
    return mejor_id if mejor_score >= 0.55 else None, mejor_score

def col_val(row, idx):
    """Lee celda de fila openpyxl de forma segura."""
    try:
        v = row[idx].value
        return v if v is not None else 0
    except IndexError:
        return 0

# ── Leer Excel ─────────────────────────────────────────────────────────────────

def leer_excel(path):
    """
    Retorna dict: { hoja: [ {nombre, horas_prog, horas_pagar, horas_virt, descuento, total} ] }
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    resultado = {}

    for hoja in HOJA_PROGRAMA:
        if hoja not in wb.sheetnames:
            print(f'  ⚠ Hoja "{hoja}" no encontrada en Excel')
            continue
        ws = wb[hoja]
        rows = list(ws.iter_rows(values_only=False))

        # Detectar fila header (buscar "NOMBRE" o "DOCENTE" en primera columna)
        header_row = 0
        for i, row in enumerate(rows):
            val = str(row[0].value or '').upper()
            if 'NOMBRE' in val or 'DOCENTE' in val:
                header_row = i
                break

        # Headers (normalizar)
        headers = [str(c.value or '').upper().strip() for c in rows[header_row]]

        # Identificar columnas clave
        def find_col(*keywords, last=False):
            """Encuentra índice de columna por keyword. last=True devuelve la última coincidencia."""
            matches = []
            for kw in keywords:
                for j, h in enumerate(headers):
                    if kw in h:
                        matches.append(j)
            if not matches:
                return None
            return matches[-1] if last else matches[0]

        col_nombre  = find_col('NOMBRE','DOCENTE') or 0
        col_noi     = find_col('NOI')
        col_hprog   = find_col('PROG')           # HORAS_PROG
        col_hpagar  = find_col('PAGAR','PAGO')   # HORAS A PAGAR / TOTAL A PAGAR
        col_hvirt   = find_col('VIRTUAL')         # HORAS VIRTUALES (monto $)
        col_desc    = find_col('DESCUENTO')
        col_subtot  = find_col('SUB')             # SUB TOTAL
        col_total   = find_col('TOTAL', last=True) # TOTAL final (última col con "TOTAL", no "SUB TOTAL")

        # Campo clínico no tiene estructura de horas
        es_cc = (hoja == 'CAMPO CLINICO')

        registros = []
        for row in rows[header_row+1:]:
            nombre = str(row[col_nombre].value or '').strip()
            if not nombre or 'TOTAL' in nombre.upper():
                continue

            def n(col):
                if col is None: return 0
                v = row[col].value
                return float(v or 0)

            if es_cc:
                # Solo total a pagar (columna 2 = índice 2)
                total = n(2)
                registros.append({
                    'nombre': nombre,
                    'horas_prog': 0,
                    'horas_pagar': 0,
                    'monto_virt': 0,
                    'descuento': 0,
                    'total': total,
                })
            else:
                hprog  = n(col_hprog)
                hpagar = n(col_hpagar) if not col_total else n(col_hprog)   # reuse
                mvirt  = n(col_hvirt)  if col_hvirt else 0
                desc   = n(col_desc)   if col_desc  else 0
                total  = n(col_total)  if col_total else n(col_hpagar)

                # Para hojas sin virtual (PREPA/ENFER/MAESTRIAS), col_hpagar ES el total
                if hoja in ('PREPA','ENFERMERIA','MAESTRIAS'):
                    hpagar = n(col_hprog)   # horas programadas
                    hreal  = n(find_col('PAGAR') or col_hpagar)
                    total  = n(find_col('TOTAL') or col_hpagar)
                    registros.append({
                        'nombre': nombre,
                        'horas_prog': hpagar,
                        'horas_pagar': hreal,
                        'monto_virt': 0,
                        'descuento': desc,
                        'total': total,
                    })
                else:
                    # Con virtual: HORAS_PROG | HORAS A PAGAR | HORAS VIRTUALES (monto) | DESCUENTO | SUB TOTAL | TOTAL
                    registros.append({
                        'nombre': nombre,
                        'horas_prog': n(col_hprog),
                        'horas_pagar': n(find_col('PAGAR')),
                        'monto_virt': n(col_hvirt) if col_hvirt else 0,
                        'descuento': desc,
                        'total': total,
                    })

        resultado[hoja] = registros
        print(f'  Excel {hoja}: {len(registros)} docentes')

    wb.close()
    return resultado

# ── Calcular nómina sistema nuevo ──────────────────────────────────────────────

def calcular_sistema(conn, docentes_db):
    """
    Para cada docente activo calcula horas presenciales (checadas),
    virtuales (evaluacion_virtual_resultado), suplencias (incidencias),
    y campo clínico — devuelve dict keyed by docente_id.
    """
    cur = conn.cursor(cursor_factory=RealDictCursor)
    resultados = {}

    for doc in docentes_db:
        did = doc['id']
        res = {
            'nombre': doc['nombre_completo'],
            'por_programa': {},   # prog_id → {hpres, hvirt, hsup, monto_virt, tarifa}
            'horas_pres_total': 0,
            'horas_virt_total': 0,
            'horas_sup_total': 0,
            'campo_clinico': 0,
        }

        # --- Presenciales por programa ---
        if doc['chec_id']:
            cur.execute("""
                WITH
                fechas AS (
                    SELECT gs::date AS fecha, EXTRACT(DOW FROM gs) AS dow
                    FROM generate_series(%s::date, %s::date, '1 day'::interval) gs
                    WHERE EXTRACT(DOW FROM gs) != 0
                ),
                checadas AS (
                    SELECT DATE(timestamp_checada) AS fecha,
                           timestamp_checada::time AS hora,
                           tipo_punch
                    FROM asistencias_checadas
                    WHERE timestamp_checada BETWEEN %s AND %s
                      AND user_id = %s
                ),
                bloques AS (
                    SELECT hc.id AS horario_id, hc.dia_semana,
                           hc.hora_inicio, hc.hora_fin, hc.horas_bloque,
                           m.programa_id, p.nombre AS programa_nombre,
                           COALESCE(a.costo_hora, p.costo_hora) AS tarifa,
                           CASE hc.dia_semana
                               WHEN 'lunes'     THEN 1 WHEN 'martes'    THEN 2
                               WHEN 'miercoles' THEN 3 WHEN 'jueves'    THEN 4
                               WHEN 'viernes'   THEN 5 WHEN 'sabado'    THEN 6
                           END AS dow_num
                    FROM horario_clases hc
                    JOIN asignaciones a ON hc.asignacion_id = a.id
                                       AND a.docente_id = %s AND a.activa = true
                    JOIN materias m     ON a.materia_id = m.id
                    JOIN programas p    ON m.programa_id = p.id
                    WHERE hc.activo = true
                )
                SELECT bf.programa_id, bf.programa_nombre, bf.tarifa, bf.horas_bloque, bf.fecha,
                    -- tipo_punch ignorado: el MB360 alterna 0/1 sin garantía de dirección
                    EXISTS(SELECT 1 FROM checadas c WHERE c.fecha=bf.fecha
                           AND c.hora BETWEEN bf.hora_inicio - INTERVAL '20 minutes'
                                          AND bf.hora_inicio + INTERVAL '10 minutes'
                    ) AS tiene_entrada,
                    EXISTS(SELECT 1 FROM checadas c WHERE c.fecha=bf.fecha
                           AND c.hora >= bf.hora_fin
                                  - (LEAST(bf.horas_bloque*10,20)||' minutes')::INTERVAL
                    ) AS tiene_salida
                FROM (SELECT b.*, f.fecha FROM bloques b JOIN fechas f ON f.dow=b.dow_num) bf
            """, (FECHA_INICIO, FECHA_FIN,   # fechas
                  FECHA_INICIO, FECHA_FIN,   # checadas
                  doc['chec_id'],            # chec_id
                  did))                      # docente_id bloques

            for b in cur.fetchall():
                pid = b['programa_id']
                if pid not in res['por_programa']:
                    res['por_programa'][pid] = {
                        'nombre': b['programa_nombre'],
                        'tarifa': float(b['tarifa'] or 0),
                        'hpres': 0, 'hvirt': 0, 'hsup': 0,
                        'monto_virt': 0,
                    }
                if b['tiene_entrada'] and b['tiene_salida']:
                    res['por_programa'][pid]['hpres'] += float(b['horas_bloque'])
                    res['horas_pres_total'] += float(b['horas_bloque'])

        # --- Virtuales ---
        cur.execute("""
            SELECT evr.horas_reales_a_pagar, evr.monto_a_pagar, evr.tarifa,
                   m.programa_id, p.nombre AS programa_nombre
            FROM evaluacion_virtual_resultado evr
            JOIN asignaciones a ON evr.asignacion_id = a.id
            JOIN materias m     ON a.materia_id = m.id
            JOIN programas p    ON m.programa_id = p.id
            WHERE evr.quincena_id = %s AND evr.docente_id = %s AND evr.aprobada = true
        """, (QUINCENA_ID, did))
        for v in cur.fetchall():
            pid = v['programa_id']
            if pid not in res['por_programa']:
                res['por_programa'][pid] = {
                    'nombre': v['programa_nombre'],
                    'tarifa': float(v['tarifa'] or 0),
                    'hpres': 0, 'hvirt': 0, 'hsup': 0, 'monto_virt': 0,
                }
            res['por_programa'][pid]['hvirt']      += float(v['horas_reales_a_pagar'] or 0)
            res['por_programa'][pid]['monto_virt'] += float(v['monto_a_pagar'] or 0)
            res['horas_virt_total'] += float(v['horas_reales_a_pagar'] or 0)

        # --- Suplencias ---
        cur.execute("""
            SELECT i.horas_suplidas, COALESCE(a.costo_hora, p.costo_hora) AS tarifa,
                   m.programa_id, p.nombre AS programa_nombre
            FROM incidencias i
            JOIN asignaciones a ON i.asignacion_id = a.id
            JOIN materias m     ON a.materia_id = m.id
            JOIN programas p    ON m.programa_id = p.id
            WHERE i.quincena_id = %s AND i.docente_suplente_id = %s
              AND i.estado = 'aprobada' AND i.tipo = 'suplencia'
              AND i.horas_suplidas IS NOT NULL
        """, (QUINCENA_ID, did))
        for s in cur.fetchall():
            pid = s['programa_id']
            if pid not in res['por_programa']:
                res['por_programa'][pid] = {
                    'nombre': s['programa_nombre'],
                    'tarifa': float(s['tarifa'] or 0),
                    'hpres': 0, 'hvirt': 0, 'hsup': 0, 'monto_virt': 0,
                }
            res['por_programa'][pid]['hsup'] += float(s['horas_suplidas'] or 0)
            res['horas_sup_total'] += float(s['horas_suplidas'] or 0)

        # --- Campo clínico (monto>0 = debe cobrar esta quincena) ---
        cur.execute("""
            SELECT monto FROM campo_clinico_quincena
            WHERE quincena_id=%s AND docente_id=%s AND COALESCE(monto,0) > 0
        """, (QUINCENA_ID, did))
        cc = cur.fetchone()
        if cc:
            res['campo_clinico'] = float(cc['monto'] or 0)

        # Solo incluir docentes con algo que cobrar
        total_horas = res['horas_pres_total'] + res['horas_virt_total'] + res['horas_sup_total']
        if total_horas > 0 or res['campo_clinico'] > 0:
            resultados[did] = res

    cur.close()
    return resultados

# ── Comparar ───────────────────────────────────────────────────────────────────

def comparar(excel_data, sistema_data, docentes_db):
    PROG_NOMBRE = {
        'PREPA': 1, 'ENFERMERIA': 2, 'NUTRICION': 3, 'LENA': 4,
        'ESPECIALIDADES': 5, 'MAESTRIAS': 6, 'CAMPO CLINICO': 7
    }
    TARIFA = {1:120, 2:140, 3:130, 4:160, 5:200, 6:220, 7:2500}

    print('\n' + '═'*110)
    print(f'{"DOCENTE":<35} {"PROG":<14} {"EXCEL_H":>8} {"SIS_H":>7} {"EXCEL_$":>10} {"SIS_$":>10} {"DIFF_$":>9} {"ESTADO"}')
    print('═'*110)

    totales = {'excel': 0, 'sistema': 0, 'match': 0, 'dif_menor': 0, 'dif_mayor': 0, 'sin_match': 0}

    for hoja, registros in excel_data.items():
        prog_id = PROG_NOMBRE.get(hoja)
        tarifa  = TARIFA.get(prog_id, 0)
        es_cc   = (hoja == 'CAMPO CLINICO')

        print(f'\n── {hoja} (tarifa ${tarifa}/h) {"─"*70}')

        for reg in registros:
            nombre_excel = reg['nombre']
            doc_id, score = buscar_docente_db(nombre_excel, docentes_db)

            # Monto Excel
            if es_cc:
                excel_monto = reg['total']
                excel_horas = 0
            else:
                excel_horas  = reg.get('horas_pagar', 0)
                excel_monto  = reg.get('total', 0)

            totales['excel'] += excel_monto

            if doc_id is None:
                print(f'  {"?? " + nombre_excel[:32]:<35} {"":14} {"":>8} {"":>7} {excel_monto:>10,.0f} {"—":>10} {"—":>9}  ❓ Sin match DB (score {score:.2f})')
                totales['sin_match'] += 1
                continue

            sis = sistema_data.get(doc_id)

            if es_cc:
                sis_monto = sis['campo_clinico'] if sis else 0
                sis_horas = 0
            else:
                if sis and prog_id in sis['por_programa']:
                    pp = sis['por_programa'][prog_id]
                    sis_horas = pp['hpres'] + pp['hvirt'] + pp['hsup']
                    sis_monto = sis_horas * tarifa + pp.get('monto_virt', 0) - (pp['hvirt'] * tarifa)
                    # monto_virt ya es el monto calculado de virtual, no horas*tarifa
                    sis_monto = pp['hpres'] * tarifa + pp['monto_virt'] + pp['hsup'] * tarifa
                else:
                    sis_horas = 0
                    sis_monto = 0

            totales['sistema'] += sis_monto
            diff = sis_monto - excel_monto

            if abs(diff) < 1:
                estado = '✅'
                totales['match'] += 1
            elif diff < 0:
                estado = f'⬇ sistema da MENOS'
                totales['dif_menor'] += 1
            else:
                estado = f'⬆ sistema da MÁS'
                totales['dif_mayor'] += 1

            nombre_db = next((d['nombre_completo'] for d in docentes_db if d['id']==doc_id), nombre_excel)
            nombre_muestra = nombre_db[:34]

            print(f'  {nombre_muestra:<35} {hoja:<14} {excel_horas:>8.1f} {sis_horas:>7.1f} {excel_monto:>10,.0f} {sis_monto:>10,.0f} {diff:>+9,.0f}  {estado}')

    print('\n' + '═'*110)
    print(f'TOTALES  Excel: ${totales["excel"]:>12,.0f}  |  Sistema: ${totales["sistema"]:>12,.0f}  |  Diff: ${totales["sistema"]-totales["excel"]:>+12,.0f}')
    print(f'Coincidencias: {totales["match"]}  |  Sistema<Excel: {totales["dif_menor"]}  |  Sistema>Excel: {totales["dif_mayor"]}  |  Sin match: {totales["sin_match"]}')
    print()

# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print('Conectando a DB...')
    conn = psycopg2.connect(DB_URL, cursor_factory=RealDictCursor)
    cur = conn.cursor(cursor_factory=RealDictCursor)

    cur.execute("SELECT id, nombre_completo, chec_id, monto_fijo_quincenal FROM docentes WHERE activo=true ORDER BY nombre_completo")
    docentes_db = cur.fetchall()
    print(f'{len(docentes_db)} docentes activos en DB')

    print('\nLeyendo Excel...')
    excel_data = leer_excel(EXCEL)

    print('\nCalculando nómina del sistema...')
    sistema_data = calcular_sistema(conn, docentes_db)
    print(f'{len(sistema_data)} docentes con horas/pago calculados')

    comparar(excel_data, sistema_data, docentes_db)

    cur.close()
    conn.close()

if __name__ == '__main__':
    main()
