"""
Importador de horarios desde reporte V1 de asistencia.
Lee el Excel generado por el sistema v1 y extrae los bloques de horario
de cada docente por programa para poblar asignaciones + horario_clases.

Uso:
    python scripts/importar_horarios_v1.py <ruta_excel> [--ciclo 2026-1] [--dry-run]

Ejemplo:
    python scripts/importar_horarios_v1.py "ASISTENCIA NOMINA 2DA MARZO.xlsx" --ciclo 2026-1
"""

import sys
import re
import unicodedata
import argparse
from collections import defaultdict
import openpyxl
import psycopg2

# ─── Mapa hoja → programa_id + modalidad ───────────────────────────────────
SHEET_PROGRAMA = {
    'PREPA':         {'programa_id': 1, 'modalidad': 'presencial'},
    'ENFERMERIA':    {'programa_id': 2, 'modalidad': 'presencial'},
    'NUTRICION':     {'programa_id': 3, 'modalidad': 'mixta'},
    'LENA':          {'programa_id': 4, 'modalidad': 'mixta'},
    'ESPECIALIDADES':{'programa_id': 5, 'modalidad': 'mixta'},
    'MAESTRIAS':     {'programa_id': 6, 'modalidad': 'virtual'},
    'CAMPO':         {'programa_id': 7, 'modalidad': 'presencial'},
}

DIA_MAP = {
    'lunes':     'lunes',
    'martes':    'martes',
    'miercoles': 'miercoles',
    'miércoles': 'miercoles',
    'miercoles': 'miercoles',
    'jueves':    'jueves',
    'viernes':   'viernes',
    'sabado':    'sabado',
    'sábado':    'sabado',
}


def normalizar(texto: str) -> str:
    """Quita tildes y convierte a minúsculas para comparación."""
    if not texto:
        return ''
    nfkd = unicodedata.normalize('NFD', texto)
    sin_tildes = ''.join(c for c in nfkd if unicodedata.category(c) != 'Mn')
    return sin_tildes.lower().strip()


def extraer_nombre_de_fila(valor: str) -> str | None:
    """
    Detecta filas tipo: 'Docente: Apellido Nombre  Adscripción: ...  Programa: ...'
    Retorna el nombre tal cual aparece (sin normalizar) para log,
    pero en el match usamos normalización.
    """
    if not isinstance(valor, str):
        return None
    m = re.match(r'Docente:\s*(.+?)\s{2,}Adscripci', valor, re.IGNORECASE)
    if not m:
        m = re.match(r'Docente:\s*(.+?)\s+Adscripci', valor, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    return None


def parse_time(t) -> str | None:
    """Convierte HH:MM a string para PostgreSQL TIME."""
    if t is None:
        return None
    if isinstance(t, str) and ':' in t:
        return t.strip()
    return None


def calcular_horas_bloque(entrada: str, salida: str) -> int:
    """Calcula horas entre dos strings HH:MM."""
    if not entrada or not salida:
        return 0
    h1, m1 = map(int, entrada.split(':'))
    h2, m2 = map(int, salida.split(':'))
    total = (h2 * 60 + m2) - (h1 * 60 + m1)
    return max(1, round(total / 60))


def leer_excel(ruta: str) -> dict:
    """
    Retorna dict: { sheet_name: { nombre_docente: { (dia, entrada, salida) } } }
    """
    wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    resultado = {}

    for sheet_name in wb.sheetnames:
        if sheet_name.upper() not in SHEET_PROGRAMA:
            print(f"  [SKIP] Hoja '{sheet_name}' no mapeada")
            continue

        ws = wb[sheet_name]
        sheet_key = sheet_name.upper()
        docentes_hoja = {}
        docente_actual = None
        leyendo_datos = False

        for row in ws.iter_rows(values_only=True):
            col0 = row[0]

            # ¿Es fila de docente?
            nombre = extraer_nombre_de_fila(str(col0) if col0 else '')
            if nombre:
                docente_actual = nombre
                leyendo_datos = False
                if docente_actual not in docentes_hoja:
                    docentes_hoja[docente_actual] = set()
                continue

            # ¿Es fila de encabezados?
            if col0 == 'FECHA' or str(col0).upper() == 'FECHA':
                leyendo_datos = True
                continue

            # ¿Fila de datos?
            if leyendo_datos and docente_actual and len(row) >= 4:
                dia_raw = row[1]
                entrada = parse_time(row[2])
                salida  = parse_time(row[3])

                if dia_raw and entrada and salida:
                    dia_norm = DIA_MAP.get(normalizar(str(dia_raw)), None)
                    if dia_norm:
                        docentes_hoja[docente_actual].add((dia_norm, entrada, salida))

        resultado[sheet_key] = docentes_hoja
        total_docentes = len(docentes_hoja)
        total_bloques = sum(len(v) for v in docentes_hoja.values())
        print(f"  [{sheet_name}] {total_docentes} docentes, {total_bloques} bloques únicos")

    return resultado


def build_nombre_index(conn) -> dict:
    """Crea índice normalizado → docente_id."""
    cur = conn.cursor()
    cur.execute("SELECT id, nombre_completo FROM docentes WHERE activo = TRUE")
    index = {}
    for row in cur.fetchall():
        doc_id, nombre = row
        index[normalizar(nombre)] = doc_id
    cur.close()
    return index


def importar(ruta: str, ciclo: str, dry_run: bool):
    print(f"\n{'[DRY RUN] ' if dry_run else ''}Importando horarios desde: {ruta}")
    print(f"Ciclo: {ciclo}\n")

    print("Leyendo Excel...")
    datos = leer_excel(ruta)

    conn = psycopg2.connect(
        host='localhost', port=5432,
        dbname='iesef_nomina',
        user='nomina_user',
        password='IESEFnomina@2026$'
    )
    cur = conn.cursor()

    nombre_index = build_nombre_index(conn)
    print(f"\n{len(nombre_index)} docentes activos en BD\n")

    # Asegurar que existe una materia genérica "Clases" por programa
    materia_index = {}  # programa_id -> materia_id
    for cfg in SHEET_PROGRAMA.values():
        pid = cfg['programa_id']
        if pid in materia_index:
            continue
        cur.execute("SELECT id FROM materias WHERE programa_id=%s AND nombre='Clases' LIMIT 1", (pid,))
        row = cur.fetchone()
        if row:
            materia_index[pid] = row[0]
        elif not dry_run:
            cur.execute(
                "INSERT INTO materias (nombre, programa_id, semestre) VALUES ('Clases', %s, 'General') RETURNING id",
                (pid,)
            )
            materia_index[pid] = cur.fetchone()[0]
        else:
            materia_index[pid] = 0  # placeholder for dry run

    stats = {'nuevas_asig': 0, 'nuevos_bloques': 0, 'ya_existen': 0, 'sin_match': 0}
    sin_match = []

    for sheet_name, docentes_hoja in datos.items():
        cfg = SHEET_PROGRAMA[sheet_name]
        programa_id = cfg['programa_id']
        modalidad   = cfg['modalidad']

        print(f"=== {sheet_name} (programa_id={programa_id}) ===")

        for nombre_excel, bloques in docentes_hoja.items():
            if not bloques:
                continue

            # Buscar docente en BD
            docente_id = nombre_index.get(normalizar(nombre_excel))
            if not docente_id:
                print(f"  [NO MATCH] '{nombre_excel}'")
                sin_match.append({'hoja': sheet_name, 'nombre': nombre_excel})
                stats['sin_match'] += 1
                continue

            # Calcular horas_semana = suma de todos los bloques únicos en la semana
            horas_semana = sum(calcular_horas_bloque(e, s) for _, e, s in bloques)

            materia_id = materia_index.get(programa_id, 0)

            # ¿Ya existe asignación para este docente+programa+ciclo?
            cur.execute(
                "SELECT id FROM asignaciones WHERE docente_id=%s AND grupo=%s AND ciclo=%s",
                (docente_id, sheet_name, ciclo)
            )
            row = cur.fetchone()

            if row:
                asig_id = row[0]
                print(f"  [EXISTS] {nombre_excel} -> asig_id={asig_id}")
                stats['ya_existen'] += 1
            else:
                # Obtener costo_hora del programa
                cur.execute("SELECT costo_hora FROM programas WHERE id=%s", (programa_id,))
                p = cur.fetchone()
                costo_hora = float(p[0]) if p else 0.0

                if not dry_run:
                    cur.execute(
                        """INSERT INTO asignaciones
                           (docente_id, materia_id, grupo, horas_semana, modalidad, costo_hora, ciclo, activa)
                           VALUES (%s, %s, %s, %s, %s, %s, %s, TRUE)
                           RETURNING id""",
                        (docente_id, materia_id, sheet_name, horas_semana, modalidad, costo_hora, ciclo)
                    )
                    asig_id = cur.fetchone()[0]
                else:
                    asig_id = 'DRY'

                print(f"  [NEW]    {nombre_excel} -> asig_id={asig_id}, {len(bloques)} bloques, {horas_semana}h/sem")
                stats['nuevas_asig'] += 1

                # Insertar horario_clases
                for (dia, entrada, salida) in sorted(bloques):
                    horas_bloque = calcular_horas_bloque(entrada, salida)
                    if not dry_run and asig_id != 'DRY':
                        cur.execute(
                            """INSERT INTO horario_clases
                               (asignacion_id, dia_semana, hora_inicio, hora_fin, horas_bloque, activo)
                               VALUES (%s, %s, %s, %s, %s, TRUE)""",
                            (asig_id, dia, entrada, salida, horas_bloque)
                        )
                    stats['nuevos_bloques'] += 1

    if not dry_run:
        conn.commit()
        print("\nOK Cambios guardados en BD")
    else:
        conn.rollback()
        print("\n[DRY RUN] No se guardo nada")

    cur.close()
    conn.close()

    print(f"\n{'-'*50}")
    print(f"  Asignaciones nuevas : {stats['nuevas_asig']}")
    print(f"  Bloques de horario  : {stats['nuevos_bloques']}")
    print(f"  Ya existían         : {stats['ya_existen']}")
    print(f"  Sin match en BD     : {stats['sin_match']}")

    if sin_match:
        print(f"\n  ATENCION - Docentes sin match ({len(sin_match)}):")
        for d in sin_match:
            print(f"     [{d['hoja']}] {d['nombre']}")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Importar horarios desde reporte V1')
    parser.add_argument('excel', help='Ruta al archivo Excel V1')
    parser.add_argument('--ciclo', default='2026-1', help='Ciclo académico (default: 2026-1)')
    parser.add_argument('--dry-run', action='store_true', help='Solo mostrar sin guardar')
    args = parser.parse_args()
    importar(args.excel, args.ciclo, args.dry_run)
